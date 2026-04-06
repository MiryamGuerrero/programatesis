from __future__ import annotations

import json
import re
import unicodedata
from itertools import product
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .rule_humanizer import build_human_rule

ROOT_DIR = Path(__file__).resolve().parents[6]
DEFAULT_XLSX = ROOT_DIR / "datosal" / "Ingredientes.xlsx"

CELL_RE = re.compile(r"^(?:'[^']+'!)?\$?([A-Za-z]{1,3})\$?\d+$")
NUMBER_RE = re.compile(r"^-?\d+(?:[\.,]\d+)?$")

FUNC_ALIASES = {
    "SI": "IF",
    "IF": "IF",
    "Y": "AND",
    "AND": "AND",
    "O": "OR",
    "OR": "OR",
    "NO": "NOT",
    "NOT": "NOT",
    "ESNUMERO": "ISNUMBER",
    "ISNUMBER": "ISNUMBER",
    "HALLAR": "FIND",
    "FIND": "FIND",
    "SEARCH": "FIND",
    "BUSCAR": "FIND",
}


def strip_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_key(text: str) -> str:
    raw = strip_text(text)
    if not raw:
        return ""
    no_accents = "".join(
        ch for ch in unicodedata.normalize("NFKD", raw) if not unicodedata.combining(ch)
    )
    lowered = no_accents.lower()
    lowered = re.sub(r"\s+", " ", lowered).strip()
    return lowered


def slugify(text: str) -> str:
    normalized = normalize_key(text)
    normalized = re.sub(r"[^a-z0-9]+", "_", normalized)
    normalized = normalized.strip("_")
    return normalized or "sin_codigo"


def detect_label_columns(headers: list[str]) -> list[str]:
    out: list[str] = []
    for header in headers:
        n = normalize_key(header)
        if "etiqueta" in n or "indice inflamatorio" in n:
            out.append(header)
    return out


def col_letter_to_index(col: str) -> int:
    value = 0
    for ch in col.upper():
        if not ("A" <= ch <= "Z"):
            return 0
        value = value * 26 + (ord(ch) - ord("A") + 1)
    return value


def _wrapped_parentheses(expr: str) -> bool:
    if not expr.startswith("(") or not expr.endswith(")"):
        return False

    depth = 0
    in_quotes = False
    for idx, ch in enumerate(expr):
        if ch == '"':
            in_quotes = not in_quotes
            continue
        if in_quotes:
            continue
        if ch == "(":
            depth += 1
        elif ch == ")":
            depth -= 1
            if depth == 0 and idx != len(expr) - 1:
                return False
    return depth == 0


def _strip_outer_parentheses(expr: str) -> str:
    out = expr.strip()
    while _wrapped_parentheses(out):
        out = out[1:-1].strip()
    return out


def _split_args(raw: str) -> list[str]:
    text = raw.strip()
    if not text:
        return []

    # Prefer ';' separator if present at top-level (Spanish Excel locale).
    separator = ";"
    if ";" not in text:
        separator = ","

    parts: list[str] = []
    start = 0
    depth = 0
    in_quotes = False

    for idx, ch in enumerate(text):
        if ch == '"':
            in_quotes = not in_quotes
            continue

        if in_quotes:
            continue

        if ch == "(":
            depth += 1
            continue
        if ch == ")":
            depth -= 1
            continue

        if depth == 0 and ch == separator:
            parts.append(text[start:idx].strip())
            start = idx + 1

    parts.append(text[start:].strip())
    return [p for p in parts if p != ""]


def _parse_function_call(expr: str) -> tuple[str, list[str]] | None:
    text = expr.strip()
    open_idx = text.find("(")
    if open_idx <= 0 or not text.endswith(")"):
        return None

    name = text[:open_idx].strip().upper()
    if not re.fullmatch(r"[A-Z_][A-Z0-9_]*", name):
        return None

    inner = text[open_idx + 1 : -1]
    args = _split_args(inner)
    return name, args


def _find_top_level_operator(expr: str) -> tuple[int, str] | None:
    # Order matters to avoid splitting '>=' as '>'.
    operators = [">=", "<=", "<>", ">", "<", "="]
    depth = 0
    in_quotes = False

    for idx, ch in enumerate(expr):
        if ch == '"':
            in_quotes = not in_quotes
            continue
        if in_quotes:
            continue
        if ch == "(":
            depth += 1
            continue
        if ch == ")":
            depth -= 1
            continue
        if depth != 0:
            continue

        for op in operators:
            if expr[idx : idx + len(op)] == op:
                return idx, op

    return None


def parse_expr(raw_expr: str) -> dict[str, Any]:
    expr = strip_text(raw_expr)
    if expr.startswith("="):
        expr = expr[1:].strip()

    expr = _strip_outer_parentheses(expr)
    if expr == "":
        return {"kind": "raw", "text": ""}

    fcall = _parse_function_call(expr)
    if fcall is not None:
        fname, args = fcall
        normalized = FUNC_ALIASES.get(fname, fname)
        return {
            "kind": "func",
            "name": normalized,
            "raw_name": fname,
            "args": [parse_expr(a) for a in args],
            "text": expr,
        }

    op_match = _find_top_level_operator(expr)
    if op_match is not None:
        pos, op = op_match
        left = expr[:pos].strip()
        right = expr[pos + len(op) :].strip()
        return {
            "kind": "compare",
            "op": op,
            "left": parse_expr(left),
            "right": parse_expr(right),
            "text": expr,
        }

    if expr.startswith('"') and expr.endswith('"') and len(expr) >= 2:
        return {"kind": "string", "value": expr[1:-1]}

    if NUMBER_RE.fullmatch(expr):
        return {"kind": "number", "value": float(expr.replace(",", "."))}

    if expr.upper() in {"VERDADERO", "TRUE"}:
        return {"kind": "bool", "value": True}
    if expr.upper() in {"FALSO", "FALSE"}:
        return {"kind": "bool", "value": False}

    cell_match = CELL_RE.fullmatch(expr)
    if cell_match:
        return {"kind": "cell", "ref": expr, "col": cell_match.group(1).upper()}

    return {"kind": "raw", "text": expr}


def _expr_to_text(node: dict[str, Any]) -> str:
    kind = node.get("kind")
    if kind == "string":
        return str(node.get("value", ""))
    if kind == "number":
        value = node.get("value")
        if value is None:
            return ""
        out = f"{float(value):.6f}".rstrip("0").rstrip(".")
        return out
    if kind == "bool":
        return "verdadero" if bool(node.get("value")) else "falso"
    if kind == "cell":
        return str(node.get("ref") or "")
    if kind == "func":
        return str(node.get("text") or "")
    if kind == "compare":
        return str(node.get("text") or "")
    return str(node.get("text") or "")


def _to_bool_ast(node: dict[str, Any]) -> dict[str, Any]:
    kind = node.get("kind")
    if kind == "func":
        fname = str(node.get("name") or "").upper()
        args = node.get("args") or []

        if fname == "AND":
            return {"type": "and", "items": [_to_bool_ast(a) for a in args]}
        if fname == "OR":
            return {"type": "or", "items": [_to_bool_ast(a) for a in args]}
        if fname == "NOT" and args:
            return {"type": "not", "item": _to_bool_ast(args[0])}

        if fname == "ISNUMBER" and args:
            first = args[0]
            if first.get("kind") == "func" and str(first.get("name") or "") == "FIND":
                fargs = first.get("args") or []
                if len(fargs) >= 2:
                    return {
                        "type": "pred",
                        "pred": {
                            "kind": "contains",
                            "needle": fargs[0],
                            "target": fargs[1],
                        },
                    }

        return {
            "type": "pred",
            "pred": {"kind": "raw", "expr": node},
        }

    if kind == "compare":
        return {
            "type": "pred",
            "pred": {
                "kind": "compare",
                "op": node.get("op"),
                "left": node.get("left"),
                "right": node.get("right"),
            },
        }

    return {
        "type": "pred",
        "pred": {"kind": "raw", "expr": node},
    }


def _merge_and(a: dict[str, Any] | None, b: dict[str, Any] | None) -> dict[str, Any] | None:
    if a is None:
        return b
    if b is None:
        return a

    items: list[dict[str, Any]] = []
    if a.get("type") == "and":
        items.extend(a.get("items") or [])
    else:
        items.append(a)

    if b.get("type") == "and":
        items.extend(b.get("items") or [])
    else:
        items.append(b)

    return {"type": "and", "items": items}


def _extract_if_branches(node: dict[str, Any]) -> list[dict[str, Any]]:
    kind = node.get("kind")
    if kind == "func" and str(node.get("name") or "") == "IF":
        args = node.get("args") or []
        if len(args) >= 3:
            cond = _to_bool_ast(args[0])
            true_branches = _extract_if_branches(args[1])
            false_branches = _extract_if_branches(args[2])

            out: list[dict[str, Any]] = []
            for branch in true_branches:
                out.append(
                    {
                        "resultado": branch["resultado"],
                        "condition": _merge_and(cond, branch.get("condition")),
                    }
                )

            neg = {"type": "not", "item": cond}
            for branch in false_branches:
                out.append(
                    {
                        "resultado": branch["resultado"],
                        "condition": _merge_and(neg, branch.get("condition")),
                    }
                )
            return out

    return [{"resultado": _expr_to_text(node), "condition": None}]


def _to_nnf(node: dict[str, Any] | None, negated: bool = False) -> dict[str, Any] | None:
    if node is None:
        return None

    ntype = node.get("type")
    if ntype == "pred":
        return {
            "type": "pred",
            "pred": node.get("pred"),
            "negated": negated,
        }

    if ntype == "not":
        return _to_nnf(node.get("item"), not negated)

    if ntype in {"and", "or"}:
        current = ntype
        if negated:
            current = "or" if ntype == "and" else "and"
        items = [_to_nnf(item, negated) for item in (node.get("items") or [])]
        items = [item for item in items if item is not None]
        if not items:
            return None
        if len(items) == 1:
            return items[0]
        return {"type": current, "items": items}

    return None


def _nnf_to_dnf(node: dict[str, Any] | None, max_terms: int = 20) -> list[list[dict[str, Any]]]:
    if node is None:
        return [[]]

    ntype = node.get("type")
    if ntype == "pred":
        return [[node]]

    if ntype == "or":
        out: list[list[dict[str, Any]]] = []
        for item in node.get("items") or []:
            out.extend(_nnf_to_dnf(item, max_terms=max_terms))
            if len(out) >= max_terms:
                return out[:max_terms]
        return out

    if ntype == "and":
        chunks = [_nnf_to_dnf(item, max_terms=max_terms) for item in (node.get("items") or [])]
        out: list[list[dict[str, Any]]] = [[]]
        for chunk in chunks:
            new_out: list[list[dict[str, Any]]] = []
            for left, right in product(out, chunk):
                merged = left + right
                new_out.append(merged)
                if len(new_out) >= max_terms:
                    break
            out = new_out[:max_terms]
            if len(out) >= max_terms:
                break
        return out

    return [[]]


def _invert_operator(op: str) -> str:
    mapping = {
        ">": "<",
        "<": ">",
        ">=": "<=",
        "<=": ">=",
        "=": "=",
        "<>": "<>",
    }
    return mapping.get(op, op)


def _is_name_like_header(header: str) -> bool:
    n = normalize_key(header)
    return any(token in n for token in ["nombre", "ingrediente", "sinonimo"])


def _literal_from_node(node: dict[str, Any]) -> tuple[str, float | str | bool | None]:
    kind = node.get("kind")
    if kind == "number":
        return "number", float(node.get("value"))
    if kind == "string":
        return "string", str(node.get("value"))
    if kind == "bool":
        return "bool", bool(node.get("value"))
    return "raw", _expr_to_text(node)


def _op_for_variable_compare(op: str) -> str:
    mapping = {
        "=": "==",
        "<>": "!=",
        ">": ">",
        "<": "<",
        ">=": ">=",
        "<=": "<=",
    }
    return mapping.get(op, op)


def _pred_to_condition(
    pred_wrapper: dict[str, Any],
    cell_to_header: dict[str, str],
    fallback_field: str,
) -> dict[str, Any] | None:
    pred = pred_wrapper.get("pred") or {}
    negated = bool(pred_wrapper.get("negated"))
    kind = pred.get("kind")

    if kind == "contains":
        needle_node = pred.get("needle") or {"kind": "raw", "text": ""}
        target_node = pred.get("target") or {"kind": "raw", "text": ""}
        needle = _expr_to_text(needle_node)

        header = None
        if target_node.get("kind") == "cell":
            col = str(target_node.get("col") or "")
            header = cell_to_header.get(col)

        if header and not _is_name_like_header(header):
            return {
                "tipo_condicion": "variable_compare",
                "variable_codigo": slugify(header),
                "operador": "contains",
                "valor_texto": needle,
                "negado": negated,
            }

        return {
            "tipo_condicion": "text_contains_name",
            "valor_texto": needle,
            "negado": negated,
        }

    if kind == "compare":
        left = pred.get("left") or {}
        right = pred.get("right") or {}
        op = str(pred.get("op") or "=")

        variable_side = None
        literal_side = None
        effective_op = op

        if left.get("kind") == "cell":
            variable_side = left
            literal_side = right
        elif right.get("kind") == "cell":
            variable_side = right
            literal_side = left
            effective_op = _invert_operator(op)

        if variable_side is None:
            return {
                "tipo_condicion": "formula_excel_result_equals",
                "campo_objetivo": fallback_field,
                "valor_texto": _expr_to_text(left) + f" {op} " + _expr_to_text(right),
                "negado": negated,
            }

        col = str(variable_side.get("col") or "")
        header = cell_to_header.get(col)
        if not header:
            return None

        lit_type, lit_value = _literal_from_node(literal_side)

        condition: dict[str, Any] = {
            "tipo_condicion": "variable_compare",
            "variable_codigo": slugify(header),
            "operador": _op_for_variable_compare(effective_op),
            "negado": negated,
        }

        if lit_type == "number":
            condition["valor_numero"] = lit_value
        else:
            condition["valor_texto"] = str(lit_value)

        return condition

    raw_expr = ""
    raw_payload = pred.get("expr")
    if isinstance(raw_payload, dict):
        raw_expr = _expr_to_text(raw_payload)

    return {
        "tipo_condicion": "formula_excel_result_equals",
        "campo_objetivo": fallback_field,
        "valor_texto": raw_expr or "condicion_no_soportada",
        "negado": negated,
    }


def formula_to_branch_rules(
    formula: str,
    label_column_name: str,
    cell_to_header: dict[str, str],
    max_dnf_terms: int = 20,
) -> list[dict[str, Any]]:
    parsed = parse_expr(formula)
    branches = _extract_if_branches(parsed)

    out: list[dict[str, Any]] = []
    for branch in branches:
        result_value = strip_text(branch.get("resultado"))
        if result_value == "":
            continue

        nnf = _to_nnf(branch.get("condition"))
        dnf_groups = _nnf_to_dnf(nnf, max_terms=max_dnf_terms)

        conditions: list[dict[str, Any]] = []
        order = 1

        if dnf_groups:
            for group_idx, group in enumerate(dnf_groups, start=1):
                for pred in group:
                    cond = _pred_to_condition(pred, cell_to_header, label_column_name)
                    if cond is None:
                        continue

                    cond["orden"] = order
                    cond["grupo_logico"] = group_idx
                    cond["conector_grupo"] = "AND"
                    conditions.append(cond)
                    order += 1

        if not conditions:
            conditions = [
                {
                    "orden": 1,
                    "grupo_logico": 1,
                    "conector_grupo": "AND",
                    "tipo_condicion": "formula_excel_result_equals",
                    "operador": "=",
                    "valor_texto": result_value,
                    "campo_objetivo": label_column_name,
                    "descripcion_humana": f"Resultado de {label_column_name} = {result_value}",
                    "negado": False,
                }
            ]

        # Complete optional keys for downstream insertions.
        normalized_conditions: list[dict[str, Any]] = []
        for c in conditions:
            normalized_conditions.append(
                {
                    "orden": int(c.get("orden", 1)),
                    "grupo_logico": int(c.get("grupo_logico", 1)),
                    "conector_grupo": str(c.get("conector_grupo", "AND")).upper(),
                    "tipo_condicion": str(c.get("tipo_condicion", "variable_compare")),
                    "variable_codigo": c.get("variable_codigo"),
                    "operador": c.get("operador"),
                    "valor_numero": c.get("valor_numero"),
                    "valor_numero_min": c.get("valor_numero_min"),
                    "valor_numero_max": c.get("valor_numero_max"),
                    "valor_texto": c.get("valor_texto"),
                    "valor_lista": c.get("valor_lista"),
                    "campo_objetivo": c.get("campo_objetivo"),
                    "negado": bool(c.get("negado", False)),
                    "descripcion_humana": c.get("descripcion_humana"),
                }
            )

        human_text = build_human_rule(
            rule_name=f"{label_column_name} -> {result_value}",
            conditions=normalized_conditions,
            result_label_name=result_value,
        )

        out.append(
            {
                "resultado": result_value,
                "condiciones": normalized_conditions,
                "expresion_humana": human_text,
            }
        )

    return out


def _extract_functions(formula: str) -> list[str]:
    found = re.findall(r"([A-Za-z_][A-Za-z0-9_]*)\s*\(", formula)
    normalized = []
    for name in found:
        n = FUNC_ALIASES.get(name.upper(), name.upper())
        if n not in normalized:
            normalized.append(n)
    return normalized


def analyze_excel_formulas(
    xlsx_path: str | Path | None = None,
    sheet_name: str = "Composicion",
    formula_row: int = 2,
    only_label_columns: bool = True,
) -> dict[str, Any]:
    target = Path(xlsx_path) if xlsx_path else DEFAULT_XLSX
    if not target.exists():
        raise FileNotFoundError(f"No existe archivo Excel: {target}")

    workbook = load_workbook(target, data_only=False)
    resolved_sheet_name = sheet_name
    if resolved_sheet_name not in workbook.sheetnames:
        wanted = normalize_key(sheet_name)
        resolved_sheet_name = ""
        for candidate in workbook.sheetnames:
            if normalize_key(candidate) == wanted:
                resolved_sheet_name = candidate
                break

    if not resolved_sheet_name:
        raise ValueError(
            f"Hoja '{sheet_name}' no encontrada. Disponibles: {', '.join(workbook.sheetnames)}"
        )

    ws = workbook[resolved_sheet_name]

    headers = [strip_text(ws.cell(1, col).value) for col in range(1, ws.max_column + 1)]
    cell_to_header: dict[str, str] = {}
    for idx, header in enumerate(headers, start=1):
        if not header:
            continue

        col = idx
        letters = []
        while col:
            col, rem = divmod(col - 1, 26)
            letters.append(chr(65 + rem))
        cell_to_header["".join(reversed(letters))] = header

    candidate_columns = headers
    if only_label_columns:
        labels = set(detect_label_columns(headers))
        candidate_columns = [h for h in headers if h in labels]

    formulas: list[dict[str, Any]] = []
    for idx, header in enumerate(headers, start=1):
        if not header:
            continue
        if header not in candidate_columns:
            continue

        raw = ws.cell(formula_row, idx).value
        if not isinstance(raw, str) or not raw.startswith("="):
            continue

        parsed_rules = formula_to_branch_rules(
            formula=raw,
            label_column_name=header,
            cell_to_header=cell_to_header,
        )

        formulas.append(
            {
                "columna": header,
                "formula_excel": raw,
                "funciones_detectadas": _extract_functions(raw),
                "reglas_detectadas": parsed_rules,
            }
        )

    return {
        "archivo": str(target),
        "hoja": resolved_sheet_name,
        "fila_formulas": formula_row,
        "columnas_totales": len([h for h in headers if h]),
        "columnas_etiqueta_detectadas": detect_label_columns(headers),
        "formulas_detectadas": formulas,
        "total_formulas_detectadas": len(formulas),
        "total_reglas_detectadas": sum(len(item["reglas_detectadas"]) for item in formulas),
    }


def serialize_rule_conditions(conditions: list[dict[str, Any]]) -> str:
    return json.dumps(conditions, ensure_ascii=False)
