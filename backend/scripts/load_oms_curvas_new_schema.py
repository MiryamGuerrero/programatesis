from __future__ import annotations

import argparse
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import psycopg
from openpyxl import load_workbook

ROOT_DIR = Path(__file__).resolve().parents[2]
DEFAULT_ENV = ROOT_DIR / "backend" / ".env"
DEFAULT_OMS_DIR = ROOT_DIR / "Fuente_datos" / "Oms datos"

PERCENTILE_CODES = [
    "p01",
    "p1",
    "p3",
    "p5",
    "p10",
    "p15",
    "p25",
    "p50",
    "p75",
    "p85",
    "p90",
    "p95",
    "p97",
    "p99",
    "p999",
]


@dataclass
class OmsCurveRow:
    indicador: str
    sexo: str
    curva_tipo: str
    edad_meses: int
    l: float | None
    m: float | None
    s: float | None
    sd5neg: float | None = None
    sd4neg: float | None = None
    sd3neg: float | None = None
    sd2neg: float | None = None
    sd1neg: float | None = None
    sd0: float | None = None
    sd1: float | None = None
    sd2: float | None = None
    sd3: float | None = None
    sd4: float | None = None
    sd5: float | None = None
    percentiles: dict[str, float] | None = None


def parse_env_file(path: Path) -> dict[str, str]:
    values: dict[str, str] = {}
    if not path.exists():
        return values

    for raw_line in path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        values[key.strip()] = value.strip().strip('"').strip("'")

    return values


def normalize_header(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    for ch in [" ", "-", "+", "_", "."]:
        text = text.replace(ch, "")
    return text


def to_float(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip()
    if not text:
        return None

    text = text.replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None


def detect_indicator(file_name: str) -> str:
    low = file_name.lower()
    if "bmi" in low:
        return "BMI"
    if "hfa" in low:
        return "HFA"
    raise ValueError(f"No se pudo identificar indicador para {file_name}")


def detect_sex(file_name: str) -> str:
    low = file_name.lower()
    if "boys" in low:
        return "M"
    if "girls" in low:
        return "F"
    raise ValueError(f"No se pudo identificar sexo para {file_name}")


def detect_curve_type(file_name: str) -> str:
    low = file_name.lower()
    if "perc" in low or "percent" in low:
        return "PERCENTIL"
    return "ZSCORE"


def find_header_map(ws) -> tuple[int, dict[str, int]]:
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=40, values_only=True), start=1):
        mapping: dict[str, int] = {}
        for idx, value in enumerate(row):
            key = normalize_header(value)
            if key and key not in mapping:
                mapping[key] = idx

        if "month" in mapping and "l" in mapping and "m" in mapping and "s" in mapping:
            return row_idx, mapping

    raise ValueError("No se encontro fila de encabezados OMS")


def parse_file(path: Path) -> list[OmsCurveRow]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    header_row, hmap = find_header_map(ws)
    indicador = detect_indicator(path.name)
    sexo = detect_sex(path.name)
    curve_type = detect_curve_type(path.name)

    rows: list[OmsCurveRow] = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        month = to_float(row[hmap["month"]])
        l = to_float(row[hmap["l"]])
        m = to_float(row[hmap["m"]])
        s = to_float(row[hmap["s"]])

        if month is None or l is None or m is None or s is None:
            continue

        edad = int(round(month))

        if curve_type == "ZSCORE":
            rows.append(
                OmsCurveRow(
                    indicador=indicador,
                    sexo=sexo,
                    curva_tipo=curve_type,
                    edad_meses=edad,
                    l=l,
                    m=m,
                    s=s,
                    sd5neg=to_float(row[hmap["sd5neg"]]) if "sd5neg" in hmap else None,
                    sd4neg=to_float(row[hmap["sd4neg"]]) if "sd4neg" in hmap else None,
                    sd3neg=to_float(row[hmap["sd3neg"]]) if "sd3neg" in hmap else None,
                    sd2neg=to_float(row[hmap["sd2neg"]]) if "sd2neg" in hmap else None,
                    sd1neg=to_float(row[hmap["sd1neg"]]) if "sd1neg" in hmap else None,
                    sd0=to_float(row[hmap["sd0"]]) if "sd0" in hmap else None,
                    sd1=to_float(row[hmap["sd1"]]) if "sd1" in hmap else None,
                    sd2=to_float(row[hmap["sd2"]]) if "sd2" in hmap else None,
                    sd3=to_float(row[hmap["sd3"]]) if "sd3" in hmap else None,
                    sd4=to_float(row[hmap["sd4"]]) if "sd4" in hmap else None,
                    sd5=to_float(row[hmap["sd5"]]) if "sd5" in hmap else None,
                )
            )
        else:
            percentiles: dict[str, float] = {}
            for code in PERCENTILE_CODES:
                if code in hmap:
                    value = to_float(row[hmap[code]])
                    if value is not None:
                        percentiles[code.upper()] = value

            if percentiles:
                rows.append(
                    OmsCurveRow(
                        indicador=indicador,
                        sexo=sexo,
                        curva_tipo=curve_type,
                        edad_meses=edad,
                        l=l,
                        m=m,
                        s=s,
                        percentiles=percentiles,
                    )
                )

    return rows


def ensure_seed_indicators(cur) -> None:
    cur.execute(
        """
        INSERT INTO referencia.indicador_antropometrico(codigo, nombre, descripcion)
        VALUES
            ('BMI', 'Indice de masa corporal para la edad', 'WHO 5-19 years'),
            ('HFA', 'Talla para la edad', 'WHO 5-19 years')
        ON CONFLICT (codigo) DO NOTHING
        """
    )


def build_lookup_ids(cur) -> tuple[dict[str, int], dict[str, int]]:
    cur.execute("SELECT codigo, id FROM referencia.indicador_antropometrico")
    indicator_ids = {row[0]: row[1] for row in cur.fetchall()}

    cur.execute("SELECT codigo, id FROM usuarios.catalogo_sexo")
    sex_ids = {row[0]: row[1] for row in cur.fetchall()}

    missing_sex = {"M", "F"} - set(sex_ids)
    if missing_sex:
        raise RuntimeError(f"Faltan codigos de sexo en usuarios.catalogo_sexo: {sorted(missing_sex)}")

    return indicator_ids, sex_ids


def upsert_curve(cur, indicador_id: int, sexo_id: int, tipo_curva: str, source_name: str) -> int:
    code = f"{indicador_id}_{sexo_id}_{tipo_curva}"
    cur.execute(
        """
        INSERT INTO referencia.oms_curva (
            codigo,
            id_indicador,
            id_sexo,
            tipo_curva,
            unidad_edad,
            fuente_archivo,
            descripcion,
            activo
        )
        VALUES (%s, %s, %s, %s, 'MESES', %s, %s, TRUE)
        ON CONFLICT (codigo)
        DO UPDATE SET
            fuente_archivo = EXCLUDED.fuente_archivo,
            descripcion = EXCLUDED.descripcion,
            activo = TRUE
        RETURNING id
        """,
        (
            code,
            indicador_id,
            sexo_id,
            tipo_curva,
            source_name,
            f"Curva OMS {tipo_curva}",
        ),
    )
    return int(cur.fetchone()[0])


def load_rows(conn: psycopg.Connection, rows_by_file: dict[Path, list[OmsCurveRow]]) -> tuple[int, int]:
    inserted_points = 0
    inserted_percentiles = 0

    with conn.cursor() as cur:
        cur.execute("SET statement_timeout = 0")
        cur.execute("SET lock_timeout = 0")
        ensure_seed_indicators(cur)
        indicator_ids, sex_ids = build_lookup_ids(cur)

        for file_path, rows in rows_by_file.items():
            if not rows:
                continue

            indicador_code = rows[0].indicador
            sexo_code = rows[0].sexo
            curve_type = rows[0].curva_tipo

            print(f"  Cargando {indicador_code} {sexo_code} ({curve_type})...")

            indicador_id = indicator_ids[indicador_code]
            sexo_id = sex_ids[sexo_code]
            curva_id = upsert_curve(cur, indicador_id, sexo_id, curve_type, file_path.name)

            if curve_type == "ZSCORE":
                payload = [
                    (
                        curva_id,
                        item.edad_meses,
                        item.l,
                        item.m,
                        item.s,
                        item.sd5neg,
                        item.sd4neg,
                        item.sd3neg,
                        item.sd2neg,
                        item.sd1neg,
                        item.sd0,
                        item.sd1,
                        item.sd2,
                        item.sd3,
                        item.sd4,
                        item.sd5,
                    )
                    for item in rows
                ]
                cur.executemany(
                    """
                    INSERT INTO referencia.oms_curva_punto (
                        id_curva, edad_valor, l, m, s,
                        sd5neg, sd4neg, sd3neg, sd2neg, sd1neg,
                        sd0, sd1, sd2, sd3, sd4, sd5
                    )
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    ON CONFLICT (id_curva, edad_valor)
                    DO UPDATE SET
                        l = EXCLUDED.l, m = EXCLUDED.m, s = EXCLUDED.s,
                        sd5neg = EXCLUDED.sd5neg, sd4neg = EXCLUDED.sd4neg,
                        sd3neg = EXCLUDED.sd3neg, sd2neg = EXCLUDED.sd2neg,
                        sd1neg = EXCLUDED.sd1neg, sd0 = EXCLUDED.sd0,
                        sd1 = EXCLUDED.sd1, sd2 = EXCLUDED.sd2,
                        sd3 = EXCLUDED.sd3, sd4 = EXCLUDED.sd4,
                        sd5 = EXCLUDED.sd5
                    """,
                    payload,
                )
                inserted_points += len(rows)
            else:
                perc_payload = []
                for item in rows:
                    if item.percentiles:
                        for perc_code, value in item.percentiles.items():
                            perc_payload.append((curva_id, item.edad_meses, perc_code, value))
                
                if perc_payload:
                    cur.executemany(
                        """
                        INSERT INTO referencia.oms_curva_percentil (
                            id_curva, edad_valor, percentil_codigo, valor
                        )
                        VALUES (%s, %s, %s, %s)
                        ON CONFLICT (id_curva, edad_valor, percentil_codigo)
                        DO UPDATE SET valor = EXCLUDED.valor
                        """,
                        perc_payload,
                    )
                    inserted_percentiles += len(perc_payload)

    conn.commit()
    return inserted_points, inserted_percentiles


def main() -> int:
    parser = argparse.ArgumentParser(description="Carga curvas OMS al esquema referencia.oms_curva*")
    parser.add_argument("--env-file", default=str(DEFAULT_ENV))
    parser.add_argument("--oms-dir", default=str(DEFAULT_OMS_DIR))
    args = parser.parse_args()

    env = parse_env_file(Path(args.env_file))
    db_url = env.get("DATABASE_URL")
    if not db_url:
        raise RuntimeError("DATABASE_URL no encontrado en el archivo .env")

    oms_dir = Path(args.oms_dir)
    if not oms_dir.exists():
        raise FileNotFoundError(f"No existe carpeta OMS: {oms_dir}")

    files = sorted(oms_dir.glob("*.xlsx"))
    if not files:
        raise FileNotFoundError(f"No hay archivos .xlsx en {oms_dir}")

    rows_by_file: dict[Path, list[OmsCurveRow]] = {}
    for file_path in files:
        print(f"Procesando archivo: {file_path.name}...")
        rows_by_file[file_path] = parse_file(file_path)
        print(f"  - {len(rows_by_file[file_path])} filas encontradas.")

    print(f"Conectando a la base de datos para cargar {len(files)} archivos...")
    with psycopg.connect(db_url, prepare_threshold=None) as conn:
        points, percentiles = load_rows(conn, rows_by_file)

    print(f"OMS_CURVAS_OK files={len(files)} zscore_rows={points} percentile_rows={percentiles}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
