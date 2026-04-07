from __future__ import annotations

import argparse
import hashlib
import json
import re
import sys
import unicodedata
from dataclasses import dataclass
from pathlib import Path
from typing import Any
from urllib.parse import unquote, urlparse

import psycopg
from openpyxl import load_workbook

ROOT_DIR = Path(__file__).resolve().parents[2]
BACKEND_DIR = ROOT_DIR / "backend"
if str(BACKEND_DIR) not in sys.path:
    sys.path.append(str(BACKEND_DIR))

from app.services.shared.cerebro.motor_etiquetas_nutricionales import (  # noqa: E402
    formula_to_branch_rules,
)

DEFAULT_XLSX = (
    ROOT_DIR / "Fuente_datos" / "Ingredientes.xlsx"
    if (ROOT_DIR / "Fuente_datos" / "Ingredientes.xlsx").exists()
    else ROOT_DIR / "datosal" / "Ingredientes.xlsx"
)
DEFAULT_ENV = ROOT_DIR / "backend" / ".env"

SHEET_COMPOSICION = "Composicion"
SHEET_DICT = "Diccionario_variables"
SHEET_UMBRALES = "Umbrales_y_formulas"

LABEL_COLUMN_HINTS = ("Etiqueta", "Indice inflamatorio", "Indice Inflamatorio")

INGREDIENT_NAME_FIELD = "Nombre"
GROUP_FIELD = "Grupo alimentario"
SUBGROUP_FIELD = "Subgrupo alimentario"
SYNONYM_FIELD = "Sinonimo"
CODE_FIELD = "Codigo"
PRICE_PER_POUND_FIELD = "Precio por libra"
PRICE_PER_GRAM_FIELD = "Precio por gramo"
POUND_IN_GRAMS = 453.592
PRICE_DECIMALS = 2
CALCULATED_DECIMALS = 4

INGREDIENT_BASE_FIELDS = {
    CODE_FIELD,
    INGREDIENT_NAME_FIELD,
    GROUP_FIELD,
    SUBGROUP_FIELD,
    SYNONYM_FIELD,
    PRICE_PER_POUND_FIELD,
    "P. comestible (por 1 g)",
}

EXCLUDED_SOURCE_FIELDS = {PRICE_PER_GRAM_FIELD}

CLASS_TO_NUTRIENT_CATEGORY = {
    "macronutrientes / energia": "MACRO",
    "perfil lipidico": "MACRO",
    "vitaminas": "VITAMINA",
    "minerales y oligoelementos": "MINERAL",
}


@dataclass
class DictVar:
    campo: str
    tipo_variable: str | None
    clasificacion: str | None
    unidad: str | None
    descripcion: str | None
    origen: str | None
    es_calculable: bool


@dataclass
class UmbralDef:
    campo_calculable: str
    que_calcula: str | None
    formula_conceptual: str | None
    columnas_origen: str | None
    unidad_salida: str | None
    criterio: str | None



def parse_env_file(path: Path) -> dict[str, str]:
    values: dict[str, str] = {}
    if not path.exists():
        return values

    for raw_line in path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        values[key.strip()] = value.strip().strip('"').strip("'")

    return values



def ensure_sslmode(database_url: str) -> str:
    if "sslmode=" in database_url:
        return database_url
    separator = "&" if "?" in database_url else "?"
    return f"{database_url}{separator}sslmode=require"


def connect_with_pooler(database_url: str, pooler_host: str, pooler_port: int) -> psycopg.Connection:
    parsed = urlparse(database_url)
    dbname = parsed.path.lstrip("/") or "postgres"
    password = unquote(parsed.password or "")

    project_ref = ""
    if parsed.hostname and parsed.hostname.startswith("db."):
        parts = parsed.hostname.split(".")
        if len(parts) > 1:
            project_ref = parts[1]

    user_candidates: list[str] = []
    if project_ref:
        user_candidates.append(f"postgres.{project_ref}")
    if parsed.username and parsed.username not in user_candidates:
        user_candidates.append(parsed.username)
    if "postgres" not in user_candidates:
        user_candidates.append("postgres")

    last_error: Exception | None = None
    for user in user_candidates:
        try:
            return psycopg.connect(
                host=pooler_host,
                port=pooler_port,
                user=user,
                password=password,
                dbname=dbname,
                sslmode="require",
                connect_timeout=30,
                prepare_threshold=None,
            )
        except Exception as exc:  # noqa: BLE001
            last_error = exc

    if last_error is not None:
        raise last_error
    raise RuntimeError("No se pudo conectar por pooler")


def connect_database(database_url: str, pooler_host: str, pooler_port: int) -> psycopg.Connection:
    try:
        return psycopg.connect(database_url, prepare_threshold=None, connect_timeout=30)
    except Exception:
        if not pooler_host:
            raise
        return connect_with_pooler(database_url, pooler_host, pooler_port)


def table_exists(cur, schema_name: str, table_name: str) -> bool:
    cur.execute("select to_regclass(%s)", (f"{schema_name}.{table_name}",))
    row = cur.fetchone()
    return bool(row and row[0])



def strip_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()



def normalize_key(text: str) -> str:
    raw = strip_text(text)
    if not raw:
        return ""
    no_accents = "".join(
        ch for ch in unicodedata.normalize("NFKD", raw) if not unicodedata.combining(ch)
    )
    lowered = no_accents.lower()
    lowered = re.sub(r"\s+", " ", lowered).strip()
    return lowered



def slugify(text: str) -> str:
    normalized = normalize_key(text)
    normalized = re.sub(r"[^a-z0-9]+", "_", normalized)
    normalized = normalized.strip("_")
    return normalized or "sin_codigo"


def bounded_slug(text: str, max_len: int) -> str:
    code = slugify(text)
    if len(code) <= max_len:
        return code

    # Keep deterministic uniqueness when truncation is needed.
    digest = hashlib.sha1(code.encode("utf-8")).hexdigest()[:8]
    head_len = max_len - 9
    if head_len < 1:
        return digest[:max_len]
    return f"{code[:head_len]}_{digest}"



def to_float(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return float(value)
    if isinstance(value, (int, float)):
        return float(value)

    text = strip_text(value)
    if text == "":
        return None

    text = text.replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None


def round_nullable(value: float | None, decimals: int) -> float | None:
    if value is None:
        return None
    return round(value, decimals)


def is_price_column(column_name: str) -> bool:
    return normalize_key(column_name) == normalize_key(PRICE_PER_POUND_FIELD)



def infer_data_type(tipo_variable: str | None, unidad: str | None) -> str:
    tv = normalize_key(tipo_variable or "")
    un = normalize_key(unidad or "")

    if "fecha" in tv:
        return "date"
    if "boole" in tv or "binaria" in tv:
        return "boolean"
    if "cualitativa" in tv or un == "texto":
        return "text"
    if "json" in tv:
        return "json"
    return "numeric"



def classify_source_state(value: Any) -> str:
    if value is None:
        return "no_reportado"

    text = strip_text(value)
    if text == "":
        return "no_reportado"

    nkey = normalize_key(text)
    if nkey in {"n/a", "na", "no aplica", "no_aplica"}:
        return "no_aplica"

    return "valor_real"



def parse_sheet_headers(ws) -> list[str]:
    headers: list[str] = []
    for col in range(1, ws.max_column + 1):
        headers.append(strip_text(ws.cell(1, col).value))
    return headers


def get_sheet(workbook, expected_name: str):
    expected = normalize_key(expected_name)
    for sheet_name in workbook.sheetnames:
        if normalize_key(sheet_name) == expected:
            return workbook[sheet_name]
    available = ", ".join(workbook.sheetnames)
    raise KeyError(f"No se encontro hoja '{expected_name}'. Hojas disponibles: {available}")



def parse_dict_variables(ws) -> dict[str, DictVar]:
    headers = parse_sheet_headers(ws)
    hmap = {normalize_key(h): idx + 1 for idx, h in enumerate(headers)}

    rows: dict[str, DictVar] = {}
    for r in range(2, ws.max_row + 1):
        campo = strip_text(ws.cell(r, hmap.get(normalize_key("Campo"), 1)).value)
        if not campo:
            continue

        tipo = strip_text(ws.cell(r, hmap.get(normalize_key("Tipo de variable"), 2)).value) or None
        clasif = strip_text(ws.cell(r, hmap.get(normalize_key("Clasificacion"), 3)).value) or None
        unidad = strip_text(ws.cell(r, hmap.get(normalize_key("Unidad"), 4)).value) or None
        descr = strip_text(ws.cell(r, hmap.get(normalize_key("Descripcion"), 5)).value) or None
        origen = strip_text(ws.cell(r, hmap.get(normalize_key("Origen"), 6)).value) or None
        es_calc_raw = strip_text(ws.cell(r, hmap.get(normalize_key("Es calculable"), 7)).value)
        es_calc = normalize_key(es_calc_raw) in {"si", "s", "yes"}

        rows[normalize_key(campo)] = DictVar(
            campo=campo,
            tipo_variable=tipo,
            clasificacion=clasif,
            unidad=unidad,
            descripcion=descr,
            origen=origen,
            es_calculable=es_calc,
        )

    return rows



def parse_umbrales(ws) -> dict[str, UmbralDef]:
    headers = parse_sheet_headers(ws)
    hmap = {normalize_key(h): idx + 1 for idx, h in enumerate(headers)}

    out: dict[str, UmbralDef] = {}
    for r in range(2, ws.max_row + 1):
        campo = strip_text(ws.cell(r, hmap.get(normalize_key("Campo calculable"), 1)).value)
        if not campo:
            continue

        out[normalize_key(campo)] = UmbralDef(
            campo_calculable=campo,
            que_calcula=strip_text(ws.cell(r, hmap.get(normalize_key("Que calcula"), 2)).value) or None,
            formula_conceptual=strip_text(
                ws.cell(r, hmap.get(normalize_key("Formula conceptual / logica"), 3)).value
            )
            or None,
            columnas_origen=strip_text(
                ws.cell(r, hmap.get(normalize_key("Columnas origen en Composicion"), 4)).value
            )
            or None,
            unidad_salida=strip_text(ws.cell(r, hmap.get(normalize_key("Unidad / salida"), 5)).value) or None,
            criterio=strip_text(
                ws.cell(r, hmap.get(normalize_key("Umbrales o criterio de interpretacion"), 6)).value
            )
            or None,
        )

    return out



def detect_label_columns(headers: list[str]) -> list[str]:
    cols: list[str] = []
    for h in headers:
        n = normalize_key(h)
        if not n:
            continue
        if "etiqueta" in n or "indice inflamatorio" in n:
            cols.append(h)
    return cols


def build_cell_to_header(headers: list[str]) -> dict[str, str]:
    mapping: dict[str, str] = {}
    for idx, header in enumerate(headers, start=1):
        if not strip_text(header):
            continue

        col = idx
        letters: list[str] = []
        while col:
            col, rem = divmod(col - 1, 26)
            letters.append(chr(65 + rem))
        mapping["".join(reversed(letters))] = header

    return mapping


def is_default_branch(conditions: list[dict[str, Any]], column_name: str, result_value: str) -> bool:
    if not conditions:
        return True

    if len(conditions) != 1:
        return False

    condition = conditions[0]
    return (
        str(condition.get("tipo_condicion") or "") == "formula_excel_result_equals"
        and strip_text(condition.get("campo_objetivo")) == strip_text(column_name)
        and strip_text(condition.get("valor_texto")) == strip_text(result_value)
    )


def branch_result_looks_unparsed(branch: dict[str, Any]) -> bool:
    result = strip_text(branch.get("resultado"))
    if not result:
        return True

    upper = result.upper()
    if upper.startswith("="):
        return True
    if "IF(" in upper or "SI(" in upper:
        return True
    if "&" in result:
        return True
    return False



def load_composition_rows(ws_values, ws_formulas) -> tuple[list[str], list[dict[str, Any]], dict[str, str | None]]:
    headers = parse_sheet_headers(ws_values)
    formulas: dict[str, str | None] = {}

    if ws_formulas.max_row >= 2:
        for col, header in enumerate(headers, start=1):
            raw = ws_formulas.cell(2, col).value
            formulas[header] = str(raw) if isinstance(raw, str) and raw.startswith("=") else None

    rows: list[dict[str, Any]] = []
    for r in range(2, ws_values.max_row + 1):
        row: dict[str, Any] = {}
        for col, header in enumerate(headers, start=1):
            row[header] = ws_values.cell(r, col).value

        if strip_text(row.get(INGREDIENT_NAME_FIELD)) == "":
            continue
        rows.append(row)

    return headers, rows, formulas



def parse_origin_columns(csv_text: str | None) -> list[str]:
    if not csv_text:
        return []
    parts = [strip_text(x) for x in csv_text.split(",")]
    return [x for x in parts if x]



def classify_nutrient_category(classification: str | None) -> str:
    key = normalize_key(classification or "")
    for token, category in CLASS_TO_NUTRIENT_CATEGORY.items():
        if token in key:
            return category

    if "vitamina" in key:
        return "VITAMINA"
    if "mineral" in key:
        return "MINERAL"
    if "macro" in key or "energia" in key:
        return "MACRO"
    return "OTRO"



def ensure_group(cur, nombre: str) -> int:
    normalized = strip_text(nombre)
    if not normalized:
        raise RuntimeError("Nombre de grupo vacio")

    cur.execute(
        """
        select id
        from dom_nutricion_catalogos.grupo_alimentario
        where lower(btrim(nombre)) = lower(btrim(%s))
        limit 1
        """,
        (normalized,),
    )
    row = cur.fetchone()
    if row:
        group_id = int(row[0])
        cur.execute(
            """
            update dom_nutricion_catalogos.grupo_alimentario
            set nombre = %s
            where id = %s
            """,
            (normalized, group_id),
        )
        return group_id

    cur.execute(
        """
        insert into dom_nutricion_catalogos.grupo_alimentario (nombre)
        values (%s)
        returning id
        """,
        (normalized,),
    )
    created = cur.fetchone()
    if not created:
        raise RuntimeError(f"No se pudo resolver grupo: {normalized}")
    return int(created[0])



def ensure_subgroup(cur, id_grupo: int, nombre: str) -> int:
    cur.execute(
        """
        select id
        from dom_nutricion_catalogos.subgrupo_alimentario
        where id_grupo_alimentario = %s
          and lower(btrim(nombre)) = lower(btrim(%s))
        limit 1
        """,
        (id_grupo, nombre),
    )
    row = cur.fetchone()
    if row:
        return int(row[0])

    cur.execute(
        """
        insert into dom_nutricion_catalogos.subgrupo_alimentario (
            id_grupo_alimentario,
            nombre,
            activo
        ) values (%s, %s, true)
        returning id
        """,
        (id_grupo, nombre),
    )
    created = cur.fetchone()
    if not created:
        raise RuntimeError(f"No se pudo crear subgrupo: {nombre}")
    return int(created[0])



def ensure_ingredient(
    cur,
    row: dict[str, Any],
    id_grupo: int | None,
    id_subgrupo: int | None,
    source_version: str,
) -> int:
    nombre = strip_text(row.get(INGREDIENT_NAME_FIELD))
    codigo = strip_text(row.get(CODE_FIELD)) or None

    precio_libra = round_nullable(to_float(row.get(PRICE_PER_POUND_FIELD)), PRICE_DECIMALS)
    parte_comestible = to_float(row.get("P. comestible (por 1 g)"))

    costo_100g = None
    if precio_libra is not None:
        costo_100g = round_nullable((precio_libra / POUND_IN_GRAMS) * 100.0, CALCULATED_DECIMALS)

    cur.execute(
        """
        select id
        from dom_nutricion_ingredientes.ingrediente
        where lower(btrim(nombre)) = lower(btrim(%s))
        limit 1
        """,
        (nombre,),
    )
    row_existing = cur.fetchone()

    if row_existing:
        ing_id = int(row_existing[0])
        cur.execute(
            """
            update dom_nutricion_ingredientes.ingrediente
            set nombre = %s,
                id_grupo_alimentario = %s,
                id_subgrupo_alimentario = %s,
                codigo_externo = coalesce(%s, codigo_externo),
                parte_comestible_factor = %s,
                precio_referencia = %s,
                costo_estimado_por_100g = %s,
                fuente_registro = 'excel',
                fecha_importacion = now(),
                version_fuente = %s,
                activo = true
            where id = %s
            """,
            (
                nombre,
                id_grupo,
                id_subgrupo,
                codigo,
                parte_comestible,
                precio_libra,
                costo_100g,
                source_version,
                ing_id,
            ),
        )
        return ing_id

    cur.execute(
        """
        insert into dom_nutricion_ingredientes.ingrediente (
            nombre,
            id_grupo_alimentario,
            id_subgrupo_alimentario,
            codigo_externo,
            parte_comestible_factor,
            precio_referencia,
            costo_estimado_por_100g,
            fuente_registro,
            fecha_importacion,
            version_fuente,
            activo
        ) values (
            %s, %s, %s, %s, %s, %s, %s, 'excel', now(), %s, true
        )
        returning id
        """,
        (
            nombre,
            id_grupo,
            id_subgrupo,
            codigo,
            parte_comestible,
            precio_libra,
            costo_100g,
            source_version,
        ),
    )
    out = cur.fetchone()
    if not out:
        raise RuntimeError(f"No se pudo crear/actualizar ingrediente: {nombre}")
    return int(out[0])



def ensure_synonym(cur, id_ingrediente: int, sinonimo: str | None) -> None:
    if not sinonimo:
        return

    normalized = strip_text(sinonimo)
    if not normalized:
        return
    if normalize_key(normalized) in {"nan", "null", "none", "n/a", "na"}:
        return

    cur.execute(
        """
        update dom_nutricion_ingredientes.ingrediente_sinonimo
        set activo = true
        where id_ingrediente = %s
          and lower(nombre_sinonimo) = lower(%s)
        """,
        (id_ingrediente, normalized),
    )
    if cur.rowcount and cur.rowcount > 0:
        return

    cur.execute(
        """
        insert into dom_nutricion_ingredientes.ingrediente_sinonimo (
            id_ingrediente,
            nombre_sinonimo,
            activo
        ) values (%s, %s, true)
        """,
        (id_ingrediente, normalized),
    )



def ensure_variable(cur, dict_var: DictVar, column_name: str) -> int:
    code = slugify(column_name)

    cur.execute(
        """
        select id
        from dom_nutricion_catalogos.variable_nutricional
        where lower(codigo) = lower(%s)
        limit 1
        """,
        (code,),
    )
    row = cur.fetchone()

    tipo_dato = infer_data_type(dict_var.tipo_variable, dict_var.unidad)

    if row:
        var_id = int(row[0])
        cur.execute(
            """
            update dom_nutricion_catalogos.variable_nutricional
            set nombre_visible = %s,
                tipo_dato = %s,
                clasificacion = %s,
                categoria_funcional = %s,
                unidad = %s,
                descripcion = %s,
                hoja_origen = %s,
                columna_origen = %s,
                origen_catalogo = 'excel',
                es_calculable = %s,
                participa_en_calculos = %s,
                participa_en_reglas = %s,
                activo = true,
                updated_at = now()
            where id = %s
            """,
            (
                dict_var.campo,
                tipo_dato,
                dict_var.clasificacion,
                dict_var.clasificacion,
                dict_var.unidad,
                dict_var.descripcion,
                SHEET_COMPOSICION,
                column_name,
                dict_var.es_calculable,
                dict_var.es_calculable,
                dict_var.es_calculable or "etiqueta" in normalize_key(column_name),
                var_id,
            ),
        )
        return var_id

    cur.execute(
        """
        insert into dom_nutricion_catalogos.variable_nutricional (
            codigo,
            nombre_visible,
            tipo_dato,
            clasificacion,
            categoria_funcional,
            unidad,
            descripcion,
            hoja_origen,
            columna_origen,
            origen_catalogo,
            es_calculable,
            participa_en_calculos,
            participa_en_reglas,
            activo
        ) values (
            %s, %s, %s, %s, %s, %s, %s,
            %s, %s, 'excel', %s, %s, %s, true
        )
        returning id
        """,
        (
            code,
            dict_var.campo,
            tipo_dato,
            dict_var.clasificacion,
            dict_var.clasificacion,
            dict_var.unidad,
            dict_var.descripcion,
            SHEET_COMPOSICION,
            column_name,
            dict_var.es_calculable,
            dict_var.es_calculable,
            dict_var.es_calculable or "etiqueta" in normalize_key(column_name),
        ),
    )
    created = cur.fetchone()
    if not created:
        raise RuntimeError(f"No se pudo crear variable: {column_name}")
    return int(created[0])



def upsert_variable_value(
    cur,
    id_ingrediente: int,
    id_variable: int,
    value: Any,
    tipo_dato: str,
    id_lote: int | None,
    source_version: str,
    column_name: str,
) -> None:
    estado = classify_source_state(value)

    valor_numerico = None
    valor_texto = None
    valor_booleano = None
    valor_fecha = None
    valor_json = None

    if estado == "valor_real":
        if tipo_dato == "numeric":
            valor_numerico = to_float(value)
            if valor_numerico is None:
                estado = "invalido"
            else:
                if is_price_column(column_name):
                    valor_numerico = round_nullable(valor_numerico, PRICE_DECIMALS)
        elif tipo_dato == "boolean":
            if isinstance(value, bool):
                valor_booleano = value
            else:
                text = normalize_key(value)
                if text in {"true", "1", "si", "yes", "s"}:
                    valor_booleano = True
                elif text in {"false", "0", "no", "n"}:
                    valor_booleano = False
                else:
                    estado = "invalido"
        elif tipo_dato == "date":
            valor_texto = strip_text(value)
            if valor_texto == "":
                estado = "no_reportado"
        elif tipo_dato == "json":
            try:
                valor_json = json.loads(strip_text(value)) if not isinstance(value, dict) else value
            except Exception:
                estado = "invalido"
        else:
            valor_texto = strip_text(value)
            if valor_texto == "":
                estado = "no_reportado"

    cur.execute(
        """
        insert into dom_nutricion_ingrediente_rel.ingrediente_variable_valor (
            id_ingrediente,
            id_variable_nutricional,
            valor_numerico,
            valor_texto,
            valor_booleano,
            valor_fecha,
            valor_json,
            estado_dato,
            origen_asignacion,
            es_fuente_primaria,
            id_importacion_lote,
            version_fuente,
            updated_at
        ) values (
            %s, %s, %s, %s, %s, %s, %s,
            %s, 'importada_desde_excel', true, %s, %s, now()
        )
        on conflict (id_ingrediente, id_variable_nutricional)
        do update set
            valor_numerico = excluded.valor_numerico,
            valor_texto = excluded.valor_texto,
            valor_booleano = excluded.valor_booleano,
            valor_fecha = excluded.valor_fecha,
            valor_json = excluded.valor_json,
            estado_dato = excluded.estado_dato,
            origen_asignacion = excluded.origen_asignacion,
            id_importacion_lote = excluded.id_importacion_lote,
            version_fuente = excluded.version_fuente,
            updated_at = now()
        """,
        (
            id_ingrediente,
            id_variable,
            valor_numerico,
            valor_texto,
            valor_booleano,
            valor_fecha,
            json.dumps(valor_json) if isinstance(valor_json, dict) else valor_json,
            estado,
            id_lote,
            source_version,
        ),
    )



def ensure_nutrient(cur, column_name: str, dict_var: DictVar) -> int:
    code = slugify(column_name).upper()
    category = classify_nutrient_category(dict_var.clasificacion)

    cur.execute(
        """
        insert into dom_nutricion_catalogos.nutriente (
            codigo,
            nombre,
            unidad_medida,
            categoria,
            activo
        ) values (%s, %s, %s, %s, true)
        on conflict (codigo)
        do update set
            nombre = excluded.nombre,
            unidad_medida = excluded.unidad_medida,
            categoria = excluded.categoria,
            activo = true
        returning id
        """,
        (code, dict_var.campo, dict_var.unidad or "g", category),
    )
    row = cur.fetchone()
    if not row:
        raise RuntimeError(f"No se pudo resolver nutriente: {column_name}")
    return int(row[0])



def upsert_ingredient_nutrient(
    cur,
    id_ingrediente: int,
    id_nutriente: int,
    value: float,
) -> None:
    cur.execute(
        """
        insert into dom_nutricion_ingrediente_rel.ingrediente_nutriente (
            id_ingrediente,
            id_nutriente,
            valor_por_100g,
            estado_dato,
            origen_asignacion,
            updated_at
        ) values (%s, %s, %s, 'valor_real', 'importada_desde_excel', now())
        on conflict (id_ingrediente, id_nutriente)
        do update set
            valor_por_100g = excluded.valor_por_100g,
            estado_dato = 'valor_real',
            origen_asignacion = 'importada_desde_excel',
            updated_at = now()
        """,
        (id_ingrediente, id_nutriente, value),
    )



def ensure_campo_derivado(
    cur,
    umbral: UmbralDef,
    formula_excel: str | None,
) -> int:
    code = slugify(umbral.campo_calculable)

    cur.execute(
        """
        select id
        from dom_nutricion_reglas.campo_derivado_definicion
        where lower(codigo) = lower(%s)
        limit 1
        """,
        (code,),
    )
    row = cur.fetchone()

    validaciones = {"criterio_interpretacion": umbral.criterio} if umbral.criterio else {}

    if row:
        out_id = int(row[0])
        cur.execute(
            """
            update dom_nutricion_reglas.campo_derivado_definicion
            set nombre_visible = %s,
                descripcion = %s,
                columnas_origen = %s,
                formula_conceptual = %s,
                formula_excel_original = %s,
                unidad = %s,
                validaciones = %s,
                activo = true,
                updated_at = now()
            where id = %s
            """,
            (
                umbral.campo_calculable,
                umbral.que_calcula,
                parse_origin_columns(umbral.columnas_origen),
                umbral.formula_conceptual or "",
                formula_excel,
                umbral.unidad_salida,
                json.dumps(validaciones),
                out_id,
            ),
        )
        return out_id

    cur.execute(
        """
        insert into dom_nutricion_reglas.campo_derivado_definicion (
            codigo,
            nombre_visible,
            descripcion,
            columnas_origen,
            formula_conceptual,
            formula_excel_original,
            unidad,
            validaciones,
            politica_dato_faltante,
            persistir_resultado,
            version,
            activo
        ) values (
            %s, %s, %s, %s, %s, %s, %s,
            %s, 'insuficiente_dato', true, 1, true
        )
        returning id
        """,
        (
            code,
            umbral.campo_calculable,
            umbral.que_calcula,
            parse_origin_columns(umbral.columnas_origen),
            umbral.formula_conceptual or "",
            formula_excel,
            umbral.unidad_salida,
            json.dumps(validaciones),
        ),
    )
    created = cur.fetchone()
    if not created:
        raise RuntimeError(f"No se pudo crear campo derivado: {umbral.campo_calculable}")
    return int(created[0])



def upsert_ingredient_derived(
    cur,
    id_ingrediente: int,
    id_campo_derivado: int,
    value: Any,
) -> None:
    estado = classify_source_state(value)

    valor_num = round_nullable(to_float(value), CALCULATED_DECIMALS)
    valor_txt = None
    if estado == "valor_real" and valor_num is None:
        valor_txt = strip_text(value)
        if valor_txt == "":
            estado = "insuficiente_dato"
    elif estado != "valor_real":
        estado = "insuficiente_dato"

    if estado == "valor_real":
        estado_calculo = "calculado"
    else:
        estado_calculo = "insuficiente_dato"

    cur.execute(
        """
        insert into dom_nutricion_ingrediente_rel.ingrediente_campo_derivado (
            id_ingrediente,
            id_campo_derivado,
            valor_numerico,
            valor_texto,
            estado_calculo,
            version_calculo,
            fecha_calculo,
            updated_at
        ) values (
            %s, %s, %s, %s, %s, 1, now(), now()
        )
        on conflict (id_ingrediente, id_campo_derivado)
        do update set
            valor_numerico = excluded.valor_numerico,
            valor_texto = excluded.valor_texto,
            estado_calculo = excluded.estado_calculo,
            version_calculo = excluded.version_calculo,
            fecha_calculo = excluded.fecha_calculo,
            updated_at = now()
        """,
        (
            id_ingrediente,
            id_campo_derivado,
            valor_num,
            valor_txt,
            estado_calculo,
        ),
    )



def ensure_label_category(cur, column_name: str) -> int:
    code = bounded_slug(column_name, 80)
    name = strip_text(column_name)

    cur.execute(
        """
        insert into dom_nutricion_catalogos.etiqueta_categoria (
            codigo,
            nombre_visible,
            descripcion,
            activa
        ) values (
            %s, %s, %s, true
        )
        on conflict (codigo)
        do update set
            nombre_visible = excluded.nombre_visible,
            descripcion = excluded.descripcion,
            activa = true,
            updated_at = now()
        returning id
        """,
        (
            code,
            name,
            f"Macro-etiqueta importada desde columna '{name}' del Excel Ingredientes.xlsx",
        ),
    )
    row = cur.fetchone()
    if not row:
        raise RuntimeError(f"No se pudo crear categoria de etiqueta: {column_name}")
    return int(row[0])


def ensure_label(
    cur,
    column_name: str,
    value_name: str,
    id_categoria: int,
    orden_resultado: int,
    es_resultado_default: bool,
) -> int:
    code = bounded_slug(f"{column_name}_{value_name}", 60)
    result_code = bounded_slug(value_name, 120)
    categoria = "Etiquetas nutricionales"
    subcategoria = column_name

    descripcion = f"Etiqueta importada desde columna '{column_name}' del Excel de ingredientes."
    objetivo = f"Clasificacion automatica desde Excel ({column_name})."
    interpretacion = f"Resultado observado en Excel: {value_name}."

    cur.execute(
        """
        select id
        from dom_nutricion_catalogos.etiqueta_nutricional
        where lower(codigo_interno) = lower(%s)
        limit 1
        """,
        (code,),
    )
    existing = cur.fetchone()

    if existing:
        label_id = int(existing[0])
        if es_resultado_default:
            cur.execute(
                """
                update dom_nutricion_catalogos.etiqueta_nutricional
                set es_resultado_default = false,
                    updated_at = now()
                where id_categoria = %s
                  and es_resultado_default
                  and id <> %s
                """,
                (id_categoria, label_id),
            )
        cur.execute(
            """
            update dom_nutricion_catalogos.etiqueta_nutricional
            set codigo_resultado = %s,
                nombre_categoria = %s,
                categoria = %s,
                subcategoria = %s,
                id_categoria = %s,
                orden_resultado = %s,
                es_resultado_default = %s,
                descripcion = %s,
                tipo_etiqueta = 'mixta',
                prioridad = 100,
                activa = true,
                objetivo_clinico = %s,
                interpretacion_base = %s,
                admite_correccion_manual = true,
                persistir_resultado = true,
                updated_at = now()
            where id = %s
            """,
            (
                result_code,
                value_name,
                categoria,
                subcategoria,
                id_categoria,
                orden_resultado,
                es_resultado_default,
                descripcion,
                objetivo,
                interpretacion,
                label_id,
            ),
        )
        return label_id

    if es_resultado_default:
        cur.execute(
            """
            update dom_nutricion_catalogos.etiqueta_nutricional
            set es_resultado_default = false,
                updated_at = now()
            where id_categoria = %s
              and es_resultado_default
            """,
            (id_categoria,),
        )

    cur.execute(
        """
        insert into dom_nutricion_catalogos.etiqueta_nutricional (
            codigo_interno,
            codigo_resultado,
            nombre_categoria,
            categoria,
            subcategoria,
            id_categoria,
            orden_resultado,
            es_resultado_default,
            descripcion,
            tipo_etiqueta,
            prioridad,
            activa,
            objetivo_clinico,
            interpretacion_base,
            admite_correccion_manual,
            persistir_resultado
        ) values (
            %s, %s, %s, %s, %s, %s, %s, %s,
            %s, 'mixta', 100, true, %s, %s, true, true
        )
        returning id
        """,
        (
            code,
            result_code,
            value_name,
            categoria,
            subcategoria,
            id_categoria,
            orden_resultado,
            es_resultado_default,
            descripcion,
            objetivo,
            interpretacion,
        ),
    )
    row = cur.fetchone()
    if not row:
        raise RuntimeError(f"No se pudo crear etiqueta: {column_name} -> {value_name}")
    return int(row[0])


def ensure_variable_by_code(cur, variable_code: str) -> int:
    normalized = strip_text(variable_code)
    if not normalized:
        raise RuntimeError("variable_code vacio")

    cur.execute(
        """
        select id
        from dom_nutricion_catalogos.variable_nutricional
        where lower(codigo) = lower(%s)
        limit 1
        """,
        (normalized,),
    )
    row = cur.fetchone()
    if row:
        return int(row[0])

    cur.execute(
        """
        insert into dom_nutricion_catalogos.variable_nutricional (
            codigo,
            nombre_visible,
            tipo_dato,
            origen_catalogo,
            participa_en_reglas,
            activo
        ) values (
            %s, %s, 'numeric', 'excel', true, true
        )
        returning id
        """,
        (normalized, normalized),
    )
    created = cur.fetchone()
    if not created:
        raise RuntimeError(f"No se pudo crear variable para codigo {variable_code}")
    return int(created[0])



def ensure_rule_for_label(
    cur,
    label_id: int,
    category_id: int,
    column_name: str,
    label_value: str,
    formula_excel: str | None,
    umbral_map: dict[str, UmbralDef],
    conditions: list[dict[str, Any]],
    priority: int,
    is_default_rule: bool,
) -> int:
    ncol = normalize_key(column_name)
    code = bounded_slug(f"excel_{slugify(column_name)}_{slugify(label_value)}", 120)

    umbral = umbral_map.get(ncol)
    campos = parse_origin_columns(umbral.columnas_origen) if umbral else []
    if not campos:
        campos = [column_name]

    umbral_resumen = {
        "fuente": "Ingredientes.xlsx",
        "columna_etiqueta": column_name,
        "valor_objetivo": label_value,
    }
    if umbral:
        umbral_resumen["criterio"] = umbral.criterio
        umbral_resumen["formula_conceptual"] = umbral.formula_conceptual

    expresion_humana = (
        f"Si se cumplen las condiciones para '{column_name}' y resultado '{label_value}', se asigna la etiqueta."
    )

    payload_conditions = conditions or []
    if not payload_conditions:
        payload_conditions = [
            {
                "orden": 1,
                "grupo_logico": 1,
                "conector_grupo": "AND",
                "tipo_condicion": "formula_excel_result_equals",
                "operador": "=",
                "valor_texto": label_value,
                "campo_objetivo": column_name,
                "negado": False,
            }
        ]

    if is_default_rule:
        cur.execute(
            """
            update dom_nutricion_reglas.etiqueta_regla_version
            set es_regla_default = false,
                updated_at = now()
            where id_categoria = %s
              and es_regla_default
              and estado = 'activa'
              and activo
              and id_etiqueta <> %s
            """,
            (category_id, label_id),
        )

    cur.execute(
        """
        insert into dom_nutricion_reglas.etiqueta_regla_version (
            id_etiqueta,
            id_categoria,
            version,
            codigo_regla,
            nombre_regla,
            estado,
            tipo_regla,
            prioridad,
            es_regla_default,
            expresion_json,
            expresion_humana,
            formula_excel_original,
            campos_intervienen,
            umbral_resumen,
            es_importada_excel,
            activo
        ) values (
            %s, %s, 1, %s, %s,
            'activa', 'automatica', %s, %s,
            %s, %s, %s, %s, %s,
            true, true
        )
        on conflict (id_etiqueta, version)
        do update set
            codigo_regla = excluded.codigo_regla,
            nombre_regla = excluded.nombre_regla,
            estado = excluded.estado,
            tipo_regla = excluded.tipo_regla,
            prioridad = excluded.prioridad,
            es_regla_default = excluded.es_regla_default,
            expresion_json = excluded.expresion_json,
            expresion_humana = excluded.expresion_humana,
            formula_excel_original = excluded.formula_excel_original,
            campos_intervienen = excluded.campos_intervienen,
            umbral_resumen = excluded.umbral_resumen,
            es_importada_excel = true,
            activo = true,
            updated_at = now()
        returning id
        """,
        (
            label_id,
            category_id,
            code,
            f"Regla importada Excel: {column_name} = {label_value}",
            priority,
            is_default_rule,
            json.dumps(
                {
                    "source": "excel",
                    "column": column_name,
                    "expected_result": label_value,
                    "formula_excel": formula_excel,
                    "conditions": payload_conditions,
                }
            ),
            expresion_humana,
            formula_excel,
            campos,
            json.dumps(umbral_resumen),
        ),
    )
    rule_row = cur.fetchone()
    if not rule_row:
        raise RuntimeError(f"No se pudo crear regla para etiqueta id={label_id}")

    rule_id = int(rule_row[0])

    cur.execute(
        """
        delete from dom_nutricion_reglas.etiqueta_regla_condicion
        where id_regla_version = %s
        """,
        (rule_id,),
    )

    for condition in payload_conditions:
        variable_code = strip_text(condition.get("variable_codigo"))
        id_variable = None
        if variable_code:
            id_variable = ensure_variable_by_code(cur, variable_code)

        cur.execute(
            """
            insert into dom_nutricion_reglas.etiqueta_regla_condicion (
                id_regla_version,
                orden,
                grupo_logico,
                conector_grupo,
                tipo_condicion,
                id_variable_nutricional,
                operador,
                valor_numero,
                valor_numero_min,
                valor_numero_max,
                valor_texto,
                valor_lista,
                campo_objetivo,
                negado,
                descripcion_humana,
                condicion_json,
                activa
            ) values (
                %s, %s, %s, %s, %s,
                %s, %s, %s, %s, %s,
                %s, %s, %s, %s, %s,
                %s::jsonb, true
            )
            """,
            (
                rule_id,
                int(condition.get("orden") or 1),
                int(condition.get("grupo_logico") or 1),
                str(condition.get("conector_grupo") or "AND").upper(),
                str(condition.get("tipo_condicion") or "variable_compare"),
                id_variable,
                condition.get("operador"),
                condition.get("valor_numero"),
                condition.get("valor_numero_min"),
                condition.get("valor_numero_max"),
                condition.get("valor_texto"),
                condition.get("valor_lista"),
                condition.get("campo_objetivo"),
                bool(condition.get("negado", False)),
                condition.get("descripcion_humana"),
                json.dumps(
                    {
                        "column": column_name,
                        "expected_result": label_value,
                        "formula_excel": formula_excel,
                        "condition": condition,
                    }
                ),
            ),
        )

    return rule_id



def upsert_ingredient_label(
    cur,
    id_ingrediente: int,
    id_etiqueta: int,
    id_regla_version: int | None,
    formula_excel: str | None,
    column_name: str,
    label_value: str,
) -> None:
    justificacion = {
        "origen": "importada_desde_excel",
        "columna_excel": column_name,
        "valor_excel": label_value,
        "formula_excel": formula_excel,
    }

    cur.execute(
        """
        insert into dom_nutricion_ingrediente_rel.ingrediente_etiqueta (
            id_ingrediente,
            id_etiqueta,
            activa,
            origen_asignacion,
            id_regla_version,
            version_regla,
            valor_justificacion,
            estado_calculo,
            es_precalculada_excel,
            manual_override,
            fecha_asignacion,
            fecha_calculo,
            updated_at
        ) values (
            %s, %s, true, 'importada_desde_excel',
            %s, 1, %s, 'calculado', true, false,
            now(), now(), now()
        )
        on conflict (id_ingrediente, id_etiqueta)
        do update set
            activa = true,
            origen_asignacion = 'importada_desde_excel',
            id_regla_version = excluded.id_regla_version,
            version_regla = excluded.version_regla,
            valor_justificacion = excluded.valor_justificacion,
            estado_calculo = 'calculado',
            es_precalculada_excel = true,
            manual_override = false,
            fecha_calculo = now(),
            updated_at = now()
        """,
        (
            id_ingrediente,
            id_etiqueta,
            id_regla_version,
            json.dumps(justificacion),
        ),
    )


def upsert_ingredient_category_result(
    cur,
    id_ingrediente: int,
    id_categoria: int,
    id_etiqueta_resultado: int,
    id_regla_version: int | None,
    prioridad_aplicada: int | None,
    detalle: dict[str, Any],
) -> None:
    cur.execute(
        """
        insert into dom_nutricion_ingrediente_rel.ingrediente_etiqueta_categoria_resultado (
            id_ingrediente,
            id_categoria,
            id_etiqueta_resultado,
            id_regla_version,
            prioridad_aplicada,
            estado_calculo,
            detalle_evaluacion,
            fecha_calculo,
            updated_at
        ) values (
            %s, %s, %s, %s, %s, 'calculado', %s::jsonb, now(), now()
        )
        on conflict (id_ingrediente, id_categoria)
        do update set
            id_etiqueta_resultado = excluded.id_etiqueta_resultado,
            id_regla_version = excluded.id_regla_version,
            prioridad_aplicada = excluded.prioridad_aplicada,
            estado_calculo = 'calculado',
            detalle_evaluacion = excluded.detalle_evaluacion,
            fecha_calculo = now(),
            updated_at = now()
        """,
        (
            id_ingrediente,
            id_categoria,
            id_etiqueta_resultado,
            id_regla_version,
            prioridad_aplicada,
            json.dumps(detalle, ensure_ascii=False),
        ),
    )



def create_import_batch(cur, source_version: str) -> int:
    cur.execute(
        """
        insert into dom_nutricion_ingrediente_rel.importacion_variable_lote (
            tipo_carga,
            archivo_nombre,
            estado,
            total_registros,
            registros_ok,
            registros_error,
            detalle_error,
            created_at,
            updated_at
        ) values (
            'excel',
            %s,
            'procesando',
            0,
            0,
            0,
            %s,
            now(),
            now()
        )
        returning id
        """,
        (str(DEFAULT_XLSX.name), json.dumps({"version_fuente": source_version})),
    )
    row = cur.fetchone()
    if not row:
        raise RuntimeError("No se pudo crear lote de importacion")
    return int(row[0])



def finalize_import_batch(
    cur,
    id_lote: int,
    total: int,
    ok: int,
    errors: int,
    detail: dict[str, Any],
) -> None:
    status = "aplicado" if errors == 0 else "error"
    cur.execute(
        """
        update dom_nutricion_ingrediente_rel.importacion_variable_lote
        set estado = %s,
            total_registros = %s,
            registros_ok = %s,
            registros_error = %s,
            detalle_error = %s,
            updated_at = now()
        where id = %s
        """,
        (
            status,
            total,
            ok,
            errors,
            json.dumps(detail),
            id_lote,
        ),
    )



def main() -> int:
    parser = argparse.ArgumentParser(
        description="Importa datos de Fuente_datos/Ingredientes.xlsx al modelo dom_* de Supabase"
    )
    parser.add_argument("--xlsx-file", default=str(DEFAULT_XLSX), help="Ruta de Ingredientes.xlsx")
    parser.add_argument("--env-file", default=str(DEFAULT_ENV), help="Ruta de archivo .env")
    parser.add_argument(
        "--source-version",
        default="Ingredientes.xlsx@2026-04-04",
        help="Version semantica de la fuente para trazabilidad",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="No hace commit, solo valida y reporta",
    )
    parser.add_argument(
        "--pooler-host",
        default="",
        help="Host pooler de Supabase (ej: aws-0-<region>.pooler.supabase.com)",
    )
    parser.add_argument(
        "--pooler-port",
        default=6543,
        type=int,
        help="Puerto del pooler de Supabase",
    )
    args = parser.parse_args()

    xlsx_file = Path(args.xlsx_file)
    env_file = Path(args.env_file)

    if not xlsx_file.exists():
        raise FileNotFoundError(f"No existe archivo: {xlsx_file}")

    env_values = parse_env_file(env_file)
    database_url = env_values.get("DATABASE_URL", "")
    if not database_url:
        raise RuntimeError("DATABASE_URL no encontrado en .env")

    database_url = ensure_sslmode(database_url)

    pooler_host = args.pooler_host or env_values.get("SUPABASE_POOLER_HOST", "")
    pooler_port = int(args.pooler_port)

    wb_values = load_workbook(xlsx_file, data_only=True)
    wb_formulas = load_workbook(xlsx_file, data_only=False)

    ws_comp_values = get_sheet(wb_values, SHEET_COMPOSICION)
    ws_comp_formulas = get_sheet(wb_formulas, SHEET_COMPOSICION)
    ws_dict = get_sheet(wb_values, SHEET_DICT)
    ws_umbrales = get_sheet(wb_values, SHEET_UMBRALES)

    dict_vars = parse_dict_variables(ws_dict)
    umbral_map = parse_umbrales(ws_umbrales)

    headers, composition_rows, formula_by_column = load_composition_rows(ws_comp_values, ws_comp_formulas)
    cell_to_header = build_cell_to_header(headers)
    label_columns = detect_label_columns(headers)

    calculable_columns: set[str] = set()
    non_calculable_columns: set[str] = set()

    for h in headers:
        dvar = dict_vars.get(normalize_key(h))
        if dvar is None:
            continue
        if dvar.es_calculable:
            calculable_columns.add(h)
        else:
            non_calculable_columns.add(h)

    # Label columns are always derivable outputs.
    for lc in label_columns:
        calculable_columns.add(lc)
        if lc in non_calculable_columns:
            non_calculable_columns.remove(lc)

    # Source fields to persist in ingrediente_variable_valor.
    source_columns = sorted(c for c in non_calculable_columns if c not in EXCLUDED_SOURCE_FIELDS)

    derived_columns = sorted(calculable_columns)

    # Nutrient source subset from non-calculable fields, excluding direct ingredient base IDs.
    nutrient_columns: set[str] = set()
    for c in source_columns:
        if c in INGREDIENT_BASE_FIELDS:
            continue
        dvar = dict_vars.get(normalize_key(c))
        if dvar is None:
            continue
        if infer_data_type(dvar.tipo_variable, dvar.unidad) != "numeric":
            continue
        ckey = normalize_key(dvar.clasificacion or "")
        if "identificacion" in ckey or "costo" in ckey:
            continue
        nutrient_columns.add(c)

    summary: dict[str, Any] = {
        "rows_total": len(composition_rows),
        "source_columns": len(source_columns),
        "derived_columns": len(derived_columns),
        "label_columns": len(label_columns),
        "nutrient_columns": len(nutrient_columns),
    }

    errors = 0
    processed = 0

    with connect_database(database_url, pooler_host, pooler_port) as conn:
        with conn.cursor() as cur:
            # Keep importer side-effect free: do not alter indexes from the loader.
            cur.execute("SET LOCAL statement_timeout = 0")
            id_lote = create_import_batch(cur, args.source_version)

            nutrient_tables_available = (
                table_exists(cur, "dom_nutricion_catalogos", "nutriente")
                and table_exists(cur, "dom_nutricion_ingrediente_rel", "ingrediente_nutriente")
            )
            if not nutrient_tables_available:
                print(
                    "WARN: tablas dom_nutricion_catalogos.nutriente / "
                    "dom_nutricion_ingrediente_rel.ingrediente_nutriente no existen; "
                    "se omite espejo legacy de nutrientes"
                )
                nutrient_columns = set()
                summary["nutrient_columns"] = 0

            variable_id_by_column: dict[str, int] = {}
            variable_type_by_column: dict[str, str] = {}
            nutrient_id_by_column: dict[str, int] = {}
            campo_derivado_id_by_column: dict[str, int] = {}

            for c in headers:
                if c in EXCLUDED_SOURCE_FIELDS:
                    continue

                dvar = dict_vars.get(normalize_key(c))
                if dvar is None:
                    continue

                var_id = ensure_variable(cur, dvar, c)
                variable_id_by_column[c] = var_id
                variable_type_by_column[c] = infer_data_type(dvar.tipo_variable, dvar.unidad)

                if c in nutrient_columns:
                    nutrient_id_by_column[c] = ensure_nutrient(cur, c, dvar)

                if c in derived_columns and c not in label_columns:
                    umbral = umbral_map.get(normalize_key(c))
                    if umbral:
                        campo_derivado_id_by_column[c] = ensure_campo_derivado(
                            cur,
                            umbral,
                            formula_by_column.get(c),
                        )

            # Build category/label/rule catalog from Excel formulas and values.
            category_id_by_column: dict[str, int] = {}
            label_id_by_column_and_value: dict[tuple[str, str], int] = {}
            rule_id_by_column_and_value: dict[tuple[str, str], int] = {}
            priority_by_column_and_value: dict[tuple[str, str], int] = {}

            for lc in label_columns:
                category_id_by_column[lc] = ensure_label_category(cur, lc)
                formula_excel = formula_by_column.get(lc)
                formula_branches: list[dict[str, Any]] = []

                if formula_excel:
                    try:
                        formula_branches = formula_to_branch_rules(
                            formula=formula_excel,
                            label_column_name=lc,
                            cell_to_header=cell_to_header,
                        )
                    except Exception as formula_exc:  # noqa: BLE001
                        print(f"WARN parse formula {lc}: {formula_exc}")

                fallback_to_values = (not formula_branches) or any(
                    branch_result_looks_unparsed(branch) for branch in formula_branches
                )

                if fallback_to_values:
                    if formula_branches:
                        print(f"WARN fallback valores observados por parse incompleto en columna: {lc}")

                    unique_values: set[str] = set()
                    for row in composition_rows:
                        value = strip_text(row.get(lc))
                        if value:
                            unique_values.add(value)

                    formula_branches = [
                        {
                            "resultado": value,
                            "condiciones": [
                                {
                                    "orden": 1,
                                    "grupo_logico": 1,
                                    "conector_grupo": "AND",
                                    "tipo_condicion": "formula_excel_result_equals",
                                    "operador": "=",
                                    "valor_texto": value,
                                    "campo_objetivo": lc,
                                    "negado": False,
                                }
                            ],
                        }
                        for value in sorted(unique_values)
                    ]

                merged_by_result: dict[str, dict[str, Any]] = {}
                for branch in formula_branches:
                    result_value = strip_text(branch.get("resultado"))
                    if not result_value:
                        continue

                    incoming_conditions = list(branch.get("condiciones") or [])
                    incoming_human = strip_text(branch.get("expresion_humana")) or None

                    current = merged_by_result.get(result_value)
                    if current is None:
                        merged_by_result[result_value] = {
                            "resultado": result_value,
                            "condiciones": incoming_conditions,
                            "expresion_humana": incoming_human,
                        }
                        continue

                    existing_conditions = current["condiciones"]
                    max_order = max((int(c.get("orden") or 0) for c in existing_conditions), default=0)
                    max_group = max((int(c.get("grupo_logico") or 0) for c in existing_conditions), default=0)

                    group_map: dict[int, int] = {}
                    for condition in incoming_conditions:
                        old_group = int(condition.get("grupo_logico") or 1)
                        if old_group not in group_map:
                            group_map[old_group] = max_group + len(group_map) + 1

                        max_order += 1
                        condition_copy = dict(condition)
                        condition_copy["orden"] = max_order
                        condition_copy["grupo_logico"] = group_map[old_group]
                        condition_copy["conector_grupo"] = "AND"
                        existing_conditions.append(condition_copy)

                    if not current.get("expresion_humana") and incoming_human:
                        current["expresion_humana"] = incoming_human

                merged_branches = list(merged_by_result.values())
                if not merged_branches:
                    continue

                default_found = False
                for idx, branch in enumerate(merged_branches, start=1):
                    label_value = branch["resultado"]
                    is_default = is_default_branch(branch.get("condiciones") or [], lc, label_value)
                    if is_default:
                        default_found = True

                    if idx == len(merged_branches) and not default_found:
                        is_default = True

                    label_id = ensure_label(
                        cur,
                        lc,
                        label_value,
                        id_categoria=category_id_by_column[lc],
                        orden_resultado=idx * 10,
                        es_resultado_default=is_default,
                    )
                    label_id_by_column_and_value[(lc, label_value)] = label_id
                    priority_by_column_and_value[(lc, label_value)] = idx

                    rule_id = ensure_rule_for_label(
                        cur,
                        label_id=label_id,
                        category_id=category_id_by_column[lc],
                        column_name=lc,
                        label_value=label_value,
                        formula_excel=formula_excel,
                        umbral_map=umbral_map,
                        conditions=branch.get("condiciones") or [],
                        priority=idx,
                        is_default_rule=is_default,
                    )
                    rule_id_by_column_and_value[(lc, label_value)] = rule_id

            # Import ingredient rows
            for row in composition_rows:
                cur.execute("savepoint sp_row_import")
                try:
                    group_name = strip_text(row.get(GROUP_FIELD))
                    subgroup_name = strip_text(row.get(SUBGROUP_FIELD))

                    id_grupo = ensure_group(cur, group_name) if group_name else None
                    id_subgrupo = (
                        ensure_subgroup(cur, id_grupo, subgroup_name)
                        if id_grupo is not None and subgroup_name
                        else None
                    )

                    id_ingrediente = ensure_ingredient(
                        cur,
                        row,
                        id_grupo=id_grupo,
                        id_subgrupo=id_subgrupo,
                        source_version=args.source_version,
                    )

                    ensure_synonym(cur, id_ingrediente, strip_text(row.get(SYNONYM_FIELD)) or None)

                    # Persist source variable values
                    for col in source_columns:
                        var_id = variable_id_by_column.get(col)
                        dtype = variable_type_by_column.get(col)
                        if var_id is None or dtype is None:
                            continue

                        upsert_variable_value(
                            cur,
                            id_ingrediente=id_ingrediente,
                            id_variable=var_id,
                            value=row.get(col),
                            tipo_dato=dtype,
                            id_lote=id_lote,
                            source_version=args.source_version,
                            column_name=col,
                        )

                    # Persist source nutrients
                    for col in nutrient_columns:
                        n_id = nutrient_id_by_column.get(col)
                        if n_id is None:
                            continue
                        val = to_float(row.get(col))
                        if val is None:
                            continue
                        upsert_ingredient_nutrient(cur, id_ingrediente, n_id, val)

                    # Persist derived numeric values as calculo cache
                    for col, campo_id in campo_derivado_id_by_column.items():
                        upsert_ingredient_derived(
                            cur,
                            id_ingrediente=id_ingrediente,
                            id_campo_derivado=campo_id,
                            value=row.get(col),
                        )

                    # Persist current label assignments imported from Excel
                    for lc in label_columns:
                        label_value = strip_text(row.get(lc))
                        if not label_value:
                            continue

                        l_id = label_id_by_column_and_value.get((lc, label_value))
                        if l_id is None:
                            continue

                        r_id = rule_id_by_column_and_value.get((lc, label_value))
                        upsert_ingredient_label(
                            cur,
                            id_ingrediente=id_ingrediente,
                            id_etiqueta=l_id,
                            id_regla_version=r_id,
                            formula_excel=formula_by_column.get(lc),
                            column_name=lc,
                            label_value=label_value,
                        )

                        c_id = category_id_by_column.get(lc)
                        if c_id is not None:
                            upsert_ingredient_category_result(
                                cur,
                                id_ingrediente=id_ingrediente,
                                id_categoria=c_id,
                                id_etiqueta_resultado=l_id,
                                id_regla_version=r_id,
                                prioridad_aplicada=priority_by_column_and_value.get((lc, label_value)),
                                detalle={
                                    "origen": "importada_desde_excel",
                                    "columna_excel": lc,
                                    "valor_excel": label_value,
                                    "formula_excel": formula_by_column.get(lc),
                                },
                            )

                    cur.execute("release savepoint sp_row_import")
                    processed += 1
                except Exception as row_exc:  # noqa: BLE001
                    cur.execute("rollback to savepoint sp_row_import")
                    cur.execute("release savepoint sp_row_import")
                    errors += 1
                    print(f"WARN fila ingrediente error: {row_exc}")

            finalize_import_batch(
                cur,
                id_lote=id_lote,
                total=len(composition_rows),
                ok=processed,
                errors=errors,
                detail=summary,
            )

        if args.dry_run:
            conn.rollback()
            print("DRY_RUN_OK")
        else:
            conn.commit()
            print("IMPORT_OK")

    print("RESUMEN", json.dumps({**summary, "procesados": processed, "errores": errors}, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
