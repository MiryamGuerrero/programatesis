from __future__ import annotations

import argparse
import csv
import time
import uuid
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

import httpx
import psycopg
from jose import jwt
from openpyxl import load_workbook

ROOT_DIR = Path(__file__).resolve().parents[2]
DATA_DIR = ROOT_DIR / "datosal"
ENV_FILE = ROOT_DIR / "backend" / ".env"
OUTPUT_CSV = ROOT_DIR / "supabase" / "oms_referencia_cleaned.csv"

INDICATOR_NAMES = {
    "IMC_EDAD": "IMC para la edad",
    "TALLA_EDAD": "Talla para la edad",
    "PESO_EDAD": "Peso para la edad",
}

SD_COLUMNS = [
    "sd4neg",
    "sd3neg",
    "sd2neg",
    "sd1neg",
    "sd0",
    "sd1",
    "sd2",
    "sd3",
    "sd4",
]


@dataclass
class OmsRow:
    indicador_codigo: str
    sexo_codigo: str
    meses: int
    l: float | None
    m: float | None
    s: float | None
    sd4neg: float | None
    sd3neg: float | None
    sd2neg: float | None
    sd1neg: float | None
    sd0: float | None
    sd1: float | None
    sd2: float | None
    sd3: float | None
    sd4: float | None
    source: str

    @property
    def key(self) -> tuple[str, str, int]:
        return (self.indicador_codigo, self.sexo_codigo, self.meses)

    @property
    def quality(self) -> int:
        # Keep rows with richer SD coverage when duplicate keys exist.
        sd_values = [
            self.sd4neg,
            self.sd3neg,
            self.sd2neg,
            self.sd1neg,
            self.sd0,
            self.sd1,
            self.sd2,
            self.sd3,
            self.sd4,
        ]
        return sum(value is not None for value in sd_values)


def parse_env_file(path: Path) -> dict[str, str]:
    values: dict[str, str] = {}
    if not path.exists():
        return values

    for raw_line in path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        values[key.strip()] = value.strip().strip('"').strip("'")
    return values


def normalize_header(cell_value: object) -> str:
    if cell_value is None:
        return ""
    value = str(cell_value).strip().lower()
    value = value.replace(" ", "")
    value = value.replace("-", "")
    value = value.replace("+", "")
    return value


def to_float(value: object) -> float | None:
    if value is None:
        return None
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        text = text.replace(",", ".")
        try:
            return round(float(text), 6)
        except ValueError:
            return None
    try:
        return round(float(value), 6)
    except (TypeError, ValueError):
        return None


def to_int_month(value: object) -> int | None:
    parsed = to_float(value)
    if parsed is None:
        return None
    return int(round(parsed))


def detect_indicator(file_name: str, sheet_name: str) -> str | None:
    source = f"{sheet_name} {file_name}".lower()
    # wfa check first because some filenames include hfa but sheet is actually wfa.
    if "wfa" in source:
        return "PESO_EDAD"
    if "hfa" in source:
        return "TALLA_EDAD"
    if "bmi" in source:
        return "IMC_EDAD"
    return None


def detect_sex(file_name: str, sheet_name: str) -> str | None:
    source = f"{sheet_name} {file_name}".lower()
    if "boys" in source:
        return "M"
    if "girls" in source:
        return "F"
    return None


def find_header(ws) -> tuple[int, dict[str, int]]:
    required = {"month", "l", "m", "s"}
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=40, values_only=True), start=1):
        normalized = [normalize_header(value) for value in row]
        mapping: dict[str, int] = {}
        for index, key in enumerate(normalized):
            if key and key not in mapping:
                mapping[key] = index
        if required.issubset(mapping.keys()):
            return row_idx, mapping
    raise ValueError("No se encontro fila de encabezados con columnas Month, L, M, S")


def rows_equal(a: OmsRow, b: OmsRow) -> bool:
    fields = ["l", "m", "s", *SD_COLUMNS]
    for field_name in fields:
        a_value = getattr(a, field_name)
        b_value = getattr(b, field_name)
        if a_value is None and b_value is None:
            continue
        if a_value is None or b_value is None:
            return False
        if abs(a_value - b_value) > 1e-8:
            return False
    return True


def parse_workbook(path: Path) -> list[OmsRow]:
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]

    indicator = detect_indicator(path.name, ws.title)
    sex = detect_sex(path.name, ws.title)
    if indicator is None or sex is None:
        raise ValueError(f"No se pudo detectar indicador/sexo en {path.name} ({ws.title})")

    header_row, columns = find_header(ws)

    rows: list[OmsRow] = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        month_value = to_int_month(row[columns["month"]])
        l_value = to_float(row[columns["l"]])
        m_value = to_float(row[columns["m"]])
        s_value = to_float(row[columns["s"]])

        if month_value is None or l_value is None or m_value is None or s_value is None:
            continue

        row_values = {
            "sd4neg": to_float(row[columns["sd4neg"]]) if "sd4neg" in columns else None,
            "sd3neg": to_float(row[columns["sd3neg"]]) if "sd3neg" in columns else None,
            "sd2neg": to_float(row[columns["sd2neg"]]) if "sd2neg" in columns else None,
            "sd1neg": to_float(row[columns["sd1neg"]]) if "sd1neg" in columns else None,
            "sd0": to_float(row[columns["sd0"]]) if "sd0" in columns else None,
            "sd1": to_float(row[columns["sd1"]]) if "sd1" in columns else None,
            "sd2": to_float(row[columns["sd2"]]) if "sd2" in columns else None,
            "sd3": to_float(row[columns["sd3"]]) if "sd3" in columns else None,
            "sd4": to_float(row[columns["sd4"]]) if "sd4" in columns else None,
        }

        rows.append(
            OmsRow(
                indicador_codigo=indicator,
                sexo_codigo=sex,
                meses=month_value,
                l=l_value,
                m=m_value,
                s=s_value,
                sd4neg=row_values["sd4neg"],
                sd3neg=row_values["sd3neg"],
                sd2neg=row_values["sd2neg"],
                sd1neg=row_values["sd1neg"],
                sd0=row_values["sd0"],
                sd1=row_values["sd1"],
                sd2=row_values["sd2"],
                sd3=row_values["sd3"],
                sd4=row_values["sd4"],
                source=f"{path.name}:{ws.title}",
            )
        )

    return rows


def collect_rows() -> tuple[list[OmsRow], dict[str, int]]:
    if not DATA_DIR.exists():
        raise FileNotFoundError(f"No existe carpeta de datos: {DATA_DIR}")

    all_rows: list[OmsRow] = []
    files = sorted(DATA_DIR.glob("*.xlsx"))
    if not files:
        raise FileNotFoundError(f"No se encontraron archivos XLSX en {DATA_DIR}")

    for file_path in files:
        all_rows.extend(parse_workbook(file_path))

    selected: dict[tuple[str, str, int], OmsRow] = {}
    replaced_lower_quality = 0
    dropped_lower_quality = 0
    conflicts = 0

    for row in all_rows:
        existing = selected.get(row.key)
        if existing is None:
            selected[row.key] = row
            continue

        if row.quality > existing.quality:
            selected[row.key] = row
            replaced_lower_quality += 1
            continue

        if row.quality < existing.quality:
            dropped_lower_quality += 1
            continue

        if not rows_equal(existing, row):
            conflicts += 1

    summary = {
        "files": len(files),
        "raw_rows": len(all_rows),
        "clean_rows": len(selected),
        "replaced_lower_quality": replaced_lower_quality,
        "dropped_lower_quality": dropped_lower_quality,
        "conflicts_same_quality": conflicts,
    }

    ordered_rows = sorted(selected.values(), key=lambda item: (item.indicador_codigo, item.sexo_codigo, item.meses))
    return ordered_rows, summary


def export_clean_csv(rows: Iterable[OmsRow]) -> None:
    OUTPUT_CSV.parent.mkdir(parents=True, exist_ok=True)
    with OUTPUT_CSV.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(
            [
                "indicador_codigo",
                "sexo_codigo",
                "meses",
                "l",
                "m",
                "s",
                "sd4neg",
                "sd3neg",
                "sd2neg",
                "sd1neg",
                "sd0",
                "sd1",
                "sd2",
                "sd3",
                "sd4",
                "source",
            ]
        )
        for row in rows:
            writer.writerow(
                [
                    row.indicador_codigo,
                    row.sexo_codigo,
                    row.meses,
                    row.l,
                    row.m,
                    row.s,
                    row.sd4neg,
                    row.sd3neg,
                    row.sd2neg,
                    row.sd1neg,
                    row.sd0,
                    row.sd1,
                    row.sd2,
                    row.sd3,
                    row.sd4,
                    row.source,
                ]
            )


def validate_table_columns(cursor) -> list[str]:
    cursor.execute(
        """
        select column_name
        from information_schema.columns
        where table_schema = 'referencia'
          and table_name = 'oms_referencia'
        """
    )
    existing = {row[0] for row in cursor.fetchall()}
    required = {
        "id_indicador",
        "id_sexo",
        "meses",
        "l",
        "m",
        "s",
        *SD_COLUMNS,
    }
    missing = sorted(required.difference(existing))
    return missing


def ensure_indicators(cursor, indicator_codes: set[str]) -> dict[str, int]:
    for indicator_code in sorted(indicator_codes):
        cursor.execute(
            """
            insert into referencia.indicador_antropometrico (codigo, nombre)
            values (%s, %s)
            on conflict (codigo) do nothing
            """,
            (indicator_code, INDICATOR_NAMES[indicator_code]),
        )

    cursor.execute(
        """
        select id, codigo
        from referencia.indicador_antropometrico
        where codigo = any(%s)
        """,
        (list(indicator_codes),),
    )
    mapping = {codigo: indicator_id for indicator_id, codigo in cursor.fetchall()}
    missing = sorted(indicator_codes.difference(mapping.keys()))
    if missing:
        raise RuntimeError(f"No se pudieron resolver indicadores: {missing}")
    return mapping


def resolve_sex_ids(cursor, sex_codes: set[str]) -> dict[str, int]:
    cursor.execute(
        """
        select id, codigo
        from usuarios.catalogo_sexo
        where codigo = any(%s)
        """,
        (list(sex_codes),),
    )
    mapping = {codigo: sex_id for sex_id, codigo in cursor.fetchall()}
    missing = sorted(sex_codes.difference(mapping.keys()))
    if missing:
        raise RuntimeError(
            "Faltan codigos en usuarios.catalogo_sexo: "
            f"{missing}. Debes cargar catalogo sexo con codigos M/F."
        )
    return mapping


def upsert_rows(connection, rows: list[OmsRow]) -> int:
    indicator_codes = {row.indicador_codigo for row in rows}
    sex_codes = {row.sexo_codigo for row in rows}

    upsert_sql = """
        insert into referencia.oms_referencia (
            id_indicador,
            id_sexo,
            meses,
            l,
            m,
            s,
            sd4neg,
            sd3neg,
            sd2neg,
            sd1neg,
            sd0,
            sd1,
            sd2,
            sd3,
            sd4
        ) values (
            %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s
        )
        on conflict (id_indicador, id_sexo, meses)
        do update set
            l = excluded.l,
            m = excluded.m,
            s = excluded.s,
            sd4neg = excluded.sd4neg,
            sd3neg = excluded.sd3neg,
            sd2neg = excluded.sd2neg,
            sd1neg = excluded.sd1neg,
            sd0 = excluded.sd0,
            sd1 = excluded.sd1,
            sd2 = excluded.sd2,
            sd3 = excluded.sd3,
            sd4 = excluded.sd4
    """

    with connection.cursor() as cursor:
        missing_columns = validate_table_columns(cursor)
        if missing_columns:
            raise RuntimeError(
                "La tabla referencia.oms_referencia no requiere rediseño completo, "
                f"pero faltan columnas: {missing_columns}"
            )

        indicator_ids = ensure_indicators(cursor, indicator_codes)
        sex_ids = resolve_sex_ids(cursor, sex_codes)

        payload = []
        for row in rows:
            payload.append(
                (
                    indicator_ids[row.indicador_codigo],
                    sex_ids[row.sexo_codigo],
                    row.meses,
                    row.l,
                    row.m,
                    row.s,
                    row.sd4neg,
                    row.sd3neg,
                    row.sd2neg,
                    row.sd1neg,
                    row.sd0,
                    row.sd1,
                    row.sd2,
                    row.sd3,
                    row.sd4,
                )
            )

        cursor.executemany(upsert_sql, payload)

    connection.commit()
    return len(rows)


def supabase_request(
    client: httpx.Client,
    *,
    supabase_url: str,
    api_key: str,
    schema: str,
    method: str,
    resource: str,
    params: dict[str, str] | None = None,
    payload: list[dict] | dict | None = None,
    prefer: str | None = None,
    bearer_token: str | None = None,
):
    headers = {
        "apikey": api_key,
        "Authorization": f"Bearer {bearer_token or api_key}",
        "Accept": "application/json",
        "Accept-Profile": schema,
    }
    if method.upper() in {"POST", "PATCH", "PUT", "DELETE"}:
        headers["Content-Profile"] = schema
    if prefer:
        headers["Prefer"] = prefer

    response = client.request(
        method=method,
        url=f"{supabase_url}/rest/v1/{resource}",
        headers=headers,
        params=params,
        json=payload,
        timeout=60.0,
    )
    if response.status_code >= 400:
        raise RuntimeError(
            f"Supabase REST fallo en {schema}.{resource} [{response.status_code}]: "
            f"{response.text[:400]}"
        )

    if not response.text.strip():
        return []
    return response.json()


def ensure_indicators_rest(
    client: httpx.Client,
    *,
    supabase_url: str,
    api_key: str,
    indicator_codes: set[str],
    bearer_token: str | None = None,
) -> dict[str, int]:
    indicator_payload = [
        {"codigo": code, "nombre": INDICATOR_NAMES[code]}
        for code in sorted(indicator_codes)
    ]

    supabase_request(
        client,
        supabase_url=supabase_url,
        api_key=api_key,
        schema="referencia",
        method="POST",
        resource="indicador_antropometrico",
        params={"on_conflict": "codigo"},
        payload=indicator_payload,
        prefer="resolution=merge-duplicates,return=minimal",
        bearer_token=bearer_token,
    )

    codes_filter = ",".join(sorted(indicator_codes))
    rows = supabase_request(
        client,
        supabase_url=supabase_url,
        api_key=api_key,
        schema="referencia",
        method="GET",
        resource="indicador_antropometrico",
        params={"select": "id,codigo", "codigo": f"in.({codes_filter})"},
        bearer_token=bearer_token,
    )
    mapping = {row["codigo"]: int(row["id"]) for row in rows}
    missing = sorted(indicator_codes.difference(mapping.keys()))
    if missing:
        raise RuntimeError(f"No se pudieron resolver indicadores via REST: {missing}")
    return mapping


def resolve_sex_ids_rest(
    client: httpx.Client,
    *,
    supabase_url: str,
    api_key: str,
    sex_codes: set[str],
    bearer_token: str | None = None,
) -> dict[str, int]:
    sex_filter = ",".join(sorted(sex_codes))
    rows = supabase_request(
        client,
        supabase_url=supabase_url,
        api_key=api_key,
        schema="usuarios",
        method="GET",
        resource="catalogo_sexo",
        params={"select": "id,codigo", "codigo": f"in.({sex_filter})"},
        bearer_token=bearer_token,
    )
    mapping = {row["codigo"]: int(row["id"]) for row in rows}
    missing = sorted(sex_codes.difference(mapping.keys()))
    if missing:
        raise RuntimeError(
            "Faltan codigos en usuarios.catalogo_sexo para REST: "
            f"{missing}. Debes cargar catalogo sexo con codigos M/F."
        )
    return mapping


def upsert_rows_via_rest(
    supabase_url: str,
    api_key: str,
    rows: list[OmsRow],
    *,
    bearer_token: str | None = None,
) -> int:
    indicator_codes = {row.indicador_codigo for row in rows}
    sex_codes = {row.sexo_codigo for row in rows}

    with httpx.Client() as client:
        indicator_ids = ensure_indicators_rest(
            client,
            supabase_url=supabase_url,
            api_key=api_key,
            indicator_codes=indicator_codes,
            bearer_token=bearer_token,
        )
        sex_ids = resolve_sex_ids_rest(
            client,
            supabase_url=supabase_url,
            api_key=api_key,
            sex_codes=sex_codes,
            bearer_token=bearer_token,
        )

        payload_rows: list[dict] = []
        for row in rows:
            payload_rows.append(
                {
                    "id_indicador": indicator_ids[row.indicador_codigo],
                    "id_sexo": sex_ids[row.sexo_codigo],
                    "meses": row.meses,
                    "l": row.l,
                    "m": row.m,
                    "s": row.s,
                    "sd4neg": row.sd4neg,
                    "sd3neg": row.sd3neg,
                    "sd2neg": row.sd2neg,
                    "sd1neg": row.sd1neg,
                    "sd0": row.sd0,
                    "sd1": row.sd1,
                    "sd2": row.sd2,
                    "sd3": row.sd3,
                    "sd4": row.sd4,
                }
            )

        chunk_size = 200
        for offset in range(0, len(payload_rows), chunk_size):
            chunk = payload_rows[offset : offset + chunk_size]
            supabase_request(
                client,
                supabase_url=supabase_url,
                api_key=api_key,
                schema="referencia",
                method="POST",
                resource="oms_referencia",
                params={"on_conflict": "id_indicador,id_sexo,meses"},
                payload=chunk,
                prefer="resolution=merge-duplicates,return=minimal",
                bearer_token=bearer_token,
            )

    return len(rows)


def build_authenticated_jwt(jwt_secret: str) -> str:
    now = int(time.time())
    claims = {
        "iss": "supabase",
        "sub": str(uuid.uuid4()),
        "role": "authenticated",
        "iat": now,
        "exp": now + 3600,
    }
    return jwt.encode(claims, jwt_secret, algorithm="HS256")


def print_summary(rows: list[OmsRow], summary: dict[str, int]) -> None:
    counts: dict[tuple[str, str], int] = {}
    for row in rows:
        key = (row.indicador_codigo, row.sexo_codigo)
        counts[key] = counts.get(key, 0) + 1

    print("Resumen de limpieza OMS")
    print(f"- Archivos analizados: {summary['files']}")
    print(f"- Filas crudas detectadas: {summary['raw_rows']}")
    print(f"- Filas limpias y unicas: {summary['clean_rows']}")
    print(f"- Duplicados reemplazados por mejor calidad: {summary['replaced_lower_quality']}")
    print(f"- Duplicados descartados por menor calidad: {summary['dropped_lower_quality']}")
    print(f"- Conflictos (misma calidad, distinto valor): {summary['conflicts_same_quality']}")

    print("Distribucion por indicador/sexo")
    for (indicator, sex), count in sorted(counts.items()):
        print(f"- {indicator} / {sex}: {count}")

    if rows:
        min_month = min(row.meses for row in rows)
        max_month = max(row.meses for row in rows)
        print(f"Rango de meses: {min_month} a {max_month}")


def main() -> None:
    parser = argparse.ArgumentParser(description="Limpia y carga datos OMS desde datosal/*.xlsx")
    parser.add_argument(
        "--apply",
        action="store_true",
        help="Inserta/actualiza datos en referencia.oms_referencia",
    )
    args = parser.parse_args()

    rows, summary = collect_rows()
    export_clean_csv(rows)
    print_summary(rows, summary)
    print(f"Archivo limpio generado: {OUTPUT_CSV}")

    if not args.apply:
        print("Modo analisis: no se hicieron cambios en base de datos.")
        return

    env_values = parse_env_file(ENV_FILE)
    database_url = env_values.get("DATABASE_URL")
    supabase_url = env_values.get("SUPABASE_URL", "").rstrip("/")
    service_role_key = env_values.get("SUPABASE_SERVICE_ROLE_KEY")
    anon_key = env_values.get("SUPABASE_ANON_KEY")
    jwt_secret = env_values.get("SUPABASE_JWT_SECRET")

    try:
        if not database_url:
            raise RuntimeError("No se encontro DATABASE_URL para conexion directa")

        with psycopg.connect(database_url) as connection:
            inserted = upsert_rows(connection, rows)
        print(f"Carga completada por PostgreSQL directo. Filas upsertadas: {inserted}")
        return
    except Exception as exc:
        print(
            "Conexion directa por PostgreSQL no disponible. "
            f"Se intentara carga por Supabase REST. Motivo: {exc}"
        )

    if not supabase_url or not service_role_key:
        print("No hay SUPABASE_URL o SUPABASE_SERVICE_ROLE_KEY para carga REST con clave de servicio.")
    else:
        try:
            inserted = upsert_rows_via_rest(supabase_url, service_role_key, rows)
            print(f"Carga completada por Supabase REST (service key). Filas upsertadas: {inserted}")
            return
        except Exception as exc:
            print(f"REST con service key fallo. Se intentara JWT authenticated. Motivo: {exc}")

    if not supabase_url or not anon_key or not jwt_secret:
        raise RuntimeError(
            "No hay SUPABASE_URL/SUPABASE_ANON_KEY/SUPABASE_JWT_SECRET para carga REST authenticated."
        )

    auth_token = build_authenticated_jwt(jwt_secret)
    inserted = upsert_rows_via_rest(
        supabase_url,
        anon_key,
        rows,
        bearer_token=auth_token,
    )
    print(f"Carga completada por Supabase REST (authenticated JWT). Filas upsertadas: {inserted}")


if __name__ == "__main__":
    main()
