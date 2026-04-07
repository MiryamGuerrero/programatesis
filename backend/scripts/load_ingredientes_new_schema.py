from __future__ import annotations

import argparse
import re
import unicodedata
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import psycopg
from openpyxl import load_workbook

ROOT_DIR = Path(__file__).resolve().parents[2]
DEFAULT_ENV = ROOT_DIR / "backend" / ".env"
DEFAULT_XLSX = ROOT_DIR / "Fuente_datos" / "Ingredientes.xlsx"

SHEET_COMPOSICION = "Composicion"
SHEET_DICCIONARIO = "Diccionario_variables"
SHEET_UMBRALES = "Umbrales_y_formulas"

BASE_NAME_COL = "Nombre"
BASE_GROUP_COL = "Grupo alimentario"
BASE_SUBGROUP_COL = "Subgrupo alimentario"
BASE_SYNONYM_COL = "Sinonimo"
BASE_CODE_COL = "Codigo"
BASE_PRICE_LB_COL = "Precio por libra"
BASE_EDIBLE_FACTOR_COL = "P. comestible (por 1 g)"

BASE_COLUMNS = {
    BASE_NAME_COL,
    BASE_GROUP_COL,
    BASE_SUBGROUP_COL,
    BASE_SYNONYM_COL,
    BASE_CODE_COL,
    BASE_PRICE_LB_COL,
    BASE_EDIBLE_FACTOR_COL,
}


@dataclass
class DictField:
    campo: str
    tipo_variable: str
    clasificacion: str
    unidad: str
    descripcion: str
    origen: str
    es_calculable: bool


@dataclass
class ThresholdField:
    campo: str
    formula_conceptual: str
    unidad_salida: str
    criterio: str


def parse_env_file(path: Path) -> dict[str, str]:
    values: dict[str, str] = {}
    if not path.exists():
        return values

    for raw_line in path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        values[key.strip()] = value.strip().strip('"').strip("'")

    return values


def strip_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_key(text: str) -> str:
    raw = strip_text(text)
    if not raw:
        return ""
    no_accents = "".join(
        ch for ch in unicodedata.normalize("NFKD", raw) if not unicodedata.combining(ch)
    )
    lowered = no_accents.lower()
    return re.sub(r"\s+", " ", lowered).strip()


def slugify(text: str, max_len: int = 80) -> str:
    value = normalize_key(text)
    value = re.sub(r"[^a-z0-9]+", "_", value).strip("_")
    if not value:
        value = "sin_codigo"
    return value[:max_len]


def to_float(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)

    text = strip_text(value)
    if not text:
        return None

    text = text.replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None


def get_sheet_by_name(workbook, expected_name: str):
    expected = normalize_key(expected_name)
    for sheet_name in workbook.sheetnames:
        if normalize_key(sheet_name) == expected:
            return workbook[sheet_name]
    raise KeyError(f"No se encontro hoja '{expected_name}'. Hojas disponibles: {workbook.sheetnames}")


def parse_headers(ws) -> list[str]:
    headers: list[str] = []
    for col in range(1, ws.max_column + 1):
        headers.append(strip_text(ws.cell(1, col).value))
    return headers


def parse_dictionary(ws) -> dict[str, DictField]:
    headers = parse_headers(ws)
    idx = {normalize_key(h): i + 1 for i, h in enumerate(headers)}

    out: dict[str, DictField] = {}
    for row in range(2, ws.max_row + 1):
        campo = strip_text(ws.cell(row, idx.get("campo", 1)).value)
        if not campo:
            continue

        tipo = strip_text(ws.cell(row, idx.get("tipo de variable", 2)).value)
        clasif = strip_text(ws.cell(row, idx.get("clasificacion", 3)).value)
        unidad = strip_text(ws.cell(row, idx.get("unidad", 4)).value)
        descr = strip_text(ws.cell(row, idx.get("descripcion", 5)).value)
        origen = strip_text(ws.cell(row, idx.get("origen", 6)).value)
        calc = normalize_key(strip_text(ws.cell(row, idx.get("es calculable", 7)).value)) in {"si", "s", "yes"}

        out[normalize_key(campo)] = DictField(
            campo=campo,
            tipo_variable=tipo,
            clasificacion=clasif,
            unidad=unidad,
            descripcion=descr,
            origen=origen,
            es_calculable=calc,
        )

    return out


def parse_thresholds(ws) -> dict[str, ThresholdField]:
    headers = parse_headers(ws)
    idx = {normalize_key(h): i + 1 for i, h in enumerate(headers)}

    out: dict[str, ThresholdField] = {}
    for row in range(2, ws.max_row + 1):
        campo = strip_text(ws.cell(row, idx.get("campo calculable", 1)).value)
        if not campo:
            continue

        out[normalize_key(campo)] = ThresholdField(
            campo=campo,
            formula_conceptual=strip_text(ws.cell(row, idx.get("formula conceptual / logica", 3)).value),
            unidad_salida=strip_text(ws.cell(row, idx.get("unidad / salida", 5)).value),
            criterio=strip_text(ws.cell(row, idx.get("umbrales o criterio de interpretacion", 6)).value),
        )

    return out


def detect_label_columns(headers: list[str]) -> list[str]:
    out: list[str] = []
    for header in headers:
        n = normalize_key(header)
        if not n:
            continue
        if "etiqueta" in n or "indice inflamatorio" in n:
            out.append(header)
    return out


def should_treat_as_text_field(field: DictField | None) -> bool:
    if field is None:
        return False
    tipo = normalize_key(field.tipo_variable)
    unidad = normalize_key(field.unidad)
    if "cualit" in tipo or "texto" in tipo:
        return True
    if unidad == "texto":
        return True
    return False


def upsert_group(cur, group_name: str) -> int | None:
    if not group_name:
        return None
    cur.execute(
        """
        INSERT INTO nutricion.grupo_alimentario(nombre)
        VALUES (%s)
        ON CONFLICT (nombre)
        DO UPDATE SET nombre = EXCLUDED.nombre
        RETURNING id
        """,
        (group_name,),
    )
    return int(cur.fetchone()[0])


def upsert_subgroup(cur, group_id: int | None, subgroup_name: str) -> int | None:
    if group_id is None or not subgroup_name:
        return None
    cur.execute(
        """
        INSERT INTO nutricion.subgrupo_alimentario(id_grupo_alimentario, nombre)
        VALUES (%s, %s)
        ON CONFLICT (id_grupo_alimentario, nombre)
        DO UPDATE SET nombre = EXCLUDED.nombre
        RETURNING id
        """,
        (group_id, subgroup_name),
    )
    return int(cur.fetchone()[0])


def upsert_ingredient(
    cur,
    code: str,
    name: str,
    group_id: int | None,
    subgroup_id: int | None,
    edible_factor: float | None,
    price_lb: float | None,
) -> int:
    price_g = None
    if price_lb is not None:
        price_g = round(price_lb / 453.592, 6)

    cur.execute(
        """
        INSERT INTO nutricion.ingrediente (
            codigo_fuente,
            nombre,
            nombre_normalizado,
            id_grupo_alimentario,
            id_subgrupo_alimentario,
            precio_libra,
            precio_gramo,
            factor_parte_comestible,
            fuente_principal,
            activo
        )
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, 'Ingredientes.xlsx', TRUE)
        ON CONFLICT (nombre)
        DO UPDATE SET
            codigo_fuente = EXCLUDED.codigo_fuente,
            nombre_normalizado = EXCLUDED.nombre_normalizado,
            id_grupo_alimentario = EXCLUDED.id_grupo_alimentario,
            id_subgrupo_alimentario = EXCLUDED.id_subgrupo_alimentario,
            precio_libra = EXCLUDED.precio_libra,
            precio_gramo = EXCLUDED.precio_gramo,
            factor_parte_comestible = EXCLUDED.factor_parte_comestible,
            fuente_principal = EXCLUDED.fuente_principal,
            activo = TRUE,
            updated_at = NOW()
        RETURNING id
        """,
        (
            code or None,
            name,
            normalize_key(name),
            group_id,
            subgroup_id,
            price_lb,
            price_g,
            edible_factor,
        ),
    )
    return int(cur.fetchone()[0])


def upsert_synonym(cur, ingredient_id: int, synonym: str) -> None:
    if not synonym:
        return
    cur.execute(
        """
        INSERT INTO nutricion.ingrediente_sinonimo(id_ingrediente, sinonimo)
        VALUES (%s, %s)
        ON CONFLICT (id_ingrediente, sinonimo) DO NOTHING
        """,
        (ingredient_id, synonym),
    )


def classify_nutrient_group(field: DictField | None) -> str:
    if field is None:
        return "OTRO"

    clasif = normalize_key(field.clasificacion)
    if "macro" in clasif or "energia" in clasif or "lipid" in clasif:
        return "MACRO"
    if "vitamina" in clasif:
        return "VITAMINA"
    if "mineral" in clasif:
        return "MINERAL"
    if "acido" in clasif or "graso" in clasif:
        return "ACIDO_GRASO"
    if "bio" in clasif or "polifen" in clasif:
        return "BIOACTIVO"
    if "probio" in clasif:
        return "PROBIOTICO"
    return "OTRO"


def ensure_nutrient(cur, header: str, field: DictField | None) -> int:
    code = slugify(header).upper()
    unit = field.unidad if field and field.unidad else "unidad"
    class_code = classify_nutrient_group(field)

    cur.execute("SELECT id FROM nutricion.clasificacion_nutriente WHERE codigo = %s", (class_code,))
    class_row = cur.fetchone()
    class_id = int(class_row[0]) if class_row else None

    cur.execute(
        """
        INSERT INTO nutricion.nutriente(
            codigo,
            nombre,
            unidad_medida,
            id_clasificacion,
            es_derivable,
            activo
        )
        VALUES (%s, %s, %s, %s, FALSE, TRUE)
        ON CONFLICT (codigo)
        DO UPDATE SET
            nombre = EXCLUDED.nombre,
            unidad_medida = EXCLUDED.unidad_medida,
            id_clasificacion = EXCLUDED.id_clasificacion,
            activo = TRUE
        RETURNING id
        """,
        (code, header, unit, class_id),
    )
    return int(cur.fetchone()[0])


def upsert_ingredient_nutrient(cur, ingredient_id: int, nutrient_id: int, value: float) -> None:
    cur.execute(
        """
        INSERT INTO nutricion.ingrediente_nutriente(
            id_ingrediente,
            id_nutriente,
            valor_por_100g,
            fuente_valor,
            version_carga,
            fecha_carga
        )
        VALUES (%s, %s, %s, 'Ingredientes.xlsx', 1, NOW())
        ON CONFLICT (id_ingrediente, id_nutriente, version_carga)
        DO UPDATE SET
            valor_por_100g = EXCLUDED.valor_por_100g,
            fuente_valor = EXCLUDED.fuente_valor,
            fecha_carga = NOW()
        """,
        (ingredient_id, nutrient_id, value),
    )


def ensure_metric(cur, header: str, field: DictField | None, thresholds: dict[str, ThresholdField]) -> int:
    code = slugify(header).upper()
    threshold = thresholds.get(normalize_key(header))
    unit = "ratio"
    formula = "NUTRIENTE:" + code
    description = "Metrica derivada desde Ingredientes.xlsx"

    if threshold is not None:
        if threshold.unidad_salida:
            unit = threshold.unidad_salida
        if threshold.formula_conceptual:
            formula = threshold.formula_conceptual
        if threshold.criterio:
            description = threshold.criterio

    if field is not None and field.unidad:
        unit = field.unidad

    cur.execute(
        """
        INSERT INTO nutricion.metrica_def(
            codigo,
            nombre,
            unidad_medida,
            formula_sql,
            descripcion,
            activa
        )
        VALUES (%s, %s, %s, %s, %s, TRUE)
        ON CONFLICT (codigo)
        DO UPDATE SET
            nombre = EXCLUDED.nombre,
            unidad_medida = EXCLUDED.unidad_medida,
            formula_sql = EXCLUDED.formula_sql,
            descripcion = EXCLUDED.descripcion,
            activa = TRUE
        RETURNING id
        """,
        (code, header, unit, formula, description),
    )
    return int(cur.fetchone()[0])


def upsert_ingredient_metric(cur, ingredient_id: int, metric_id: int, value: float) -> None:
    cur.execute(
        """
        INSERT INTO nutricion.ingrediente_metrica(
            id_ingrediente,
            id_metrica,
            valor_numerico,
            version_calculo,
            fecha_calculo
        )
        VALUES (%s, %s, %s, 1, NOW())
        ON CONFLICT (id_ingrediente, id_metrica, version_calculo)
        DO UPDATE SET
            valor_numerico = EXCLUDED.valor_numerico,
            fecha_calculo = NOW()
        """,
        (ingredient_id, metric_id, value),
    )


def ensure_label_catalog(cur, label_header: str) -> tuple[int, int]:
    label_code = slugify(label_header).upper()
    etq_code = f"ET_{label_code}"

    cur.execute(
        """
        INSERT INTO nutricion.etiqueta_nutricional(codigo, nombre_visible)
        VALUES (%s, %s)
        ON CONFLICT (codigo)
        DO UPDATE SET nombre_visible = EXCLUDED.nombre_visible
        RETURNING id
        """,
        (label_code, label_header),
    )
    nutrition_label_id = int(cur.fetchone()[0])

    cur.execute("SELECT id FROM etiquetado.tipo_regla WHERE codigo = 'COMPUESTA' LIMIT 1")
    tipo = cur.fetchone()
    tipo_id = int(tipo[0]) if tipo else None

    cur.execute(
        """
        INSERT INTO etiquetado.etiqueta(codigo, nombre_visible, descripcion, id_tipo_regla_base, activa)
        VALUES (%s, %s, %s, %s, TRUE)
        ON CONFLICT (codigo)
        DO UPDATE SET
            nombre_visible = EXCLUDED.nombre_visible,
            descripcion = EXCLUDED.descripcion,
            id_tipo_regla_base = COALESCE(EXCLUDED.id_tipo_regla_base, etiquetado.etiqueta.id_tipo_regla_base),
            activa = TRUE
        RETURNING id
        """,
        (etq_code, label_header, f"Importado desde columna {label_header}", tipo_id),
    )
    etiqueta_id = int(cur.fetchone()[0])

    return nutrition_label_id, etiqueta_id


def ensure_rule_version(cur, etiqueta_id: int) -> int:
    cur.execute(
        """
        INSERT INTO etiquetado.regla_version(id_etiqueta, version_numero, estado, observacion, fecha_publicacion)
        VALUES (%s, 1, 'PUBLICADA', 'Version importada desde Ingredientes.xlsx', NOW())
        ON CONFLICT (id_etiqueta, version_numero)
        DO UPDATE SET
            estado = 'PUBLICADA',
            observacion = EXCLUDED.observacion,
            fecha_publicacion = NOW()
        RETURNING id
        """,
        (etiqueta_id,),
    )
    return int(cur.fetchone()[0])


def ensure_sub_label(cur, etiqueta_id: int, regla_version_id: int, value_text: str) -> int:
    code = slugify(value_text).upper()

    cur.execute(
        """
        SELECT COALESCE(MAX(prioridad), 0) + 1
        FROM etiquetado.subetiqueta
        WHERE id_etiqueta = %s
        """,
        (etiqueta_id,),
    )
    priority = int(cur.fetchone()[0])

    cur.execute(
        """
        INSERT INTO etiquetado.subetiqueta(id_etiqueta, codigo, nombre_visible, prioridad, activa)
        VALUES (%s, %s, %s, %s, TRUE)
        ON CONFLICT (id_etiqueta, codigo)
        DO UPDATE SET
            nombre_visible = EXCLUDED.nombre_visible,
            activa = TRUE
        RETURNING id
        """,
        (etiqueta_id, code, value_text, priority),
    )
    sub_id = int(cur.fetchone()[0])

    cur.execute(
        """
        INSERT INTO etiquetado.subetiqueta_regla(
            id_regla_version,
            id_subetiqueta,
            prioridad_evaluacion,
            tipo_evaluacion,
            observacion
        )
        VALUES (%s, %s, %s, 'FIRST_MATCH', 'Importado desde Ingredientes.xlsx')
        ON CONFLICT (id_regla_version, id_subetiqueta)
        DO UPDATE SET prioridad_evaluacion = EXCLUDED.prioridad_evaluacion
        """,
        (regla_version_id, sub_id, priority),
    )

    return sub_id


def insert_result_label(
    cur,
    ingredient_id: int,
    nutrition_label_id: int,
    etiqueta_id: int,
    sub_id: int,
    regla_version_id: int,
    value_text: str,
) -> None:
    cur.execute(
        """
        INSERT INTO nutricion.ingrediente_etiqueta(id_ingrediente, id_etiqueta)
        VALUES (%s, %s)
        ON CONFLICT (id_ingrediente, id_etiqueta) DO NOTHING
        """,
        (ingredient_id, nutrition_label_id),
    )

    cur.execute(
        """
        INSERT INTO etiquetado.ingrediente_resultado_etiqueta(
            id_ingrediente,
            id_etiqueta,
            id_subetiqueta,
            id_regla_version,
            valor_disparador,
            detalle_evaluacion,
            fecha_calculo
        )
        VALUES (
            %s,
            %s,
            %s,
            %s,
            NULL,
            jsonb_build_object('valor_texto', %s, 'origen', 'Ingredientes.xlsx'),
            NOW()
        )
        ON CONFLICT (id_ingrediente, id_etiqueta, id_regla_version)
        DO UPDATE SET
            id_subetiqueta = EXCLUDED.id_subetiqueta,
            detalle_evaluacion = EXCLUDED.detalle_evaluacion,
            fecha_calculo = NOW()
        """,
        (
            ingredient_id,
            etiqueta_id,
            sub_id,
            regla_version_id,
            value_text,
        ),
    )


def main() -> int:
    parser = argparse.ArgumentParser(description="Carga Ingredientes.xlsx al nuevo esquema")
    parser.add_argument("--env-file", default=str(DEFAULT_ENV))
    parser.add_argument("--xlsx", default=str(DEFAULT_XLSX))
    args = parser.parse_args()

    env = parse_env_file(Path(args.env_file))
    db_url = env.get("DATABASE_URL")
    if not db_url:
        raise RuntimeError("DATABASE_URL no encontrado")

    xlsx = Path(args.xlsx)
    if not xlsx.exists():
        raise FileNotFoundError(f"No existe archivo: {xlsx}")

    wb = load_workbook(xlsx, data_only=True)
    ws_comp = get_sheet_by_name(wb, SHEET_COMPOSICION)
    ws_dict = get_sheet_by_name(wb, SHEET_DICCIONARIO)
    ws_umbral = get_sheet_by_name(wb, SHEET_UMBRALES)

    headers = parse_headers(ws_comp)
    dict_fields = parse_dictionary(ws_dict)
    threshold_fields = parse_thresholds(ws_umbral)
    label_columns = detect_label_columns(headers)

    header_index = {h: idx + 1 for idx, h in enumerate(headers) if h}

    ingredient_by_row: dict[int, int] = {}
    nutrient_cache: dict[str, int] = {}
    metric_cache: dict[str, int] = {}
    label_cache: dict[str, tuple[int, int, int]] = {}
    sublabel_cache: dict[tuple[int, str], int] = {}

    inserted_ingredients = 0
    inserted_nutrients = 0
    inserted_metrics = 0
    inserted_label_results = 0

    with psycopg.connect(db_url, prepare_threshold=None) as conn:
        with conn.cursor() as cur:
            cur.execute("SET statement_timeout = 0")
            cur.execute("SET lock_timeout = 0")
        with conn.cursor() as cur:
            for row in range(2, ws_comp.max_row + 1):
                name = strip_text(ws_comp.cell(row, header_index.get(BASE_NAME_COL, 1)).value)
                if not name:
                    continue

                group_name = strip_text(ws_comp.cell(row, header_index.get(BASE_GROUP_COL, 2)).value)
                subgroup_name = strip_text(ws_comp.cell(row, header_index.get(BASE_SUBGROUP_COL, 3)).value)
                synonym = strip_text(ws_comp.cell(row, header_index.get(BASE_SYNONYM_COL, 4)).value)
                code = strip_text(ws_comp.cell(row, header_index.get(BASE_CODE_COL, 5)).value)
                price_lb = to_float(ws_comp.cell(row, header_index.get(BASE_PRICE_LB_COL, 6)).value)
                edible = to_float(ws_comp.cell(row, header_index.get(BASE_EDIBLE_FACTOR_COL, 7)).value)

                group_id = upsert_group(cur, group_name)
                subgroup_id = upsert_subgroup(cur, group_id, subgroup_name)
                ingredient_id = upsert_ingredient(cur, code, name, group_id, subgroup_id, edible, price_lb)
                upsert_synonym(cur, ingredient_id, synonym)
                ingredient_by_row[row] = ingredient_id
                inserted_ingredients += 1

            for header in headers:
                if not header or header in BASE_COLUMNS or header in label_columns:
                    continue

                fkey = normalize_key(header)
                fmeta = dict_fields.get(fkey)

                if should_treat_as_text_field(fmeta):
                    continue

                is_metric = fmeta.es_calculable if fmeta else False

                if is_metric:
                    metric_id = ensure_metric(cur, header, fmeta, threshold_fields)
                    metric_cache[header] = metric_id
                else:
                    nutrient_id = ensure_nutrient(cur, header, fmeta)
                    nutrient_cache[header] = nutrient_id

            for row, ingredient_id in ingredient_by_row.items():
                for header, nutrient_id in nutrient_cache.items():
                    value = to_float(ws_comp.cell(row, header_index[header]).value)
                    if value is None:
                        continue
                    upsert_ingredient_nutrient(cur, ingredient_id, nutrient_id, value)
                    inserted_nutrients += 1

                for header, metric_id in metric_cache.items():
                    value = to_float(ws_comp.cell(row, header_index[header]).value)
                    if value is None:
                        continue
                    upsert_ingredient_metric(cur, ingredient_id, metric_id, value)
                    inserted_metrics += 1

                for label_header in label_columns:
                    label_value = strip_text(ws_comp.cell(row, header_index[label_header]).value)
                    if not label_value:
                        continue

                    if label_header not in label_cache:
                        nutrition_label_id, etiqueta_id = ensure_label_catalog(cur, label_header)
                        regla_version_id = ensure_rule_version(cur, etiqueta_id)
                        label_cache[label_header] = (nutrition_label_id, etiqueta_id, regla_version_id)

                    nutrition_label_id, etiqueta_id, regla_version_id = label_cache[label_header]

                    sub_key = (etiqueta_id, label_value)
                    if sub_key not in sublabel_cache:
                        sublabel_cache[sub_key] = ensure_sub_label(cur, etiqueta_id, regla_version_id, label_value)

                    sub_id = sublabel_cache[sub_key]
                    insert_result_label(
                        cur,
                        ingredient_id,
                        nutrition_label_id,
                        etiqueta_id,
                        sub_id,
                        regla_version_id,
                        label_value,
                    )
                    inserted_label_results += 1

        conn.commit()

    print(
        "INGREDIENTES_ETL_OK "
        f"ingredientes={inserted_ingredients} "
        f"nutrientes={len(nutrient_cache)} valores_nutrientes={inserted_nutrients} "
        f"metricas={len(metric_cache)} valores_metricas={inserted_metrics} "
        f"resultados_etiqueta={inserted_label_results}"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
