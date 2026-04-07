from __future__ import annotations

import json
import re
import sys
import unicodedata
from pathlib import Path
from urllib.parse import unquote, urlparse

import psycopg
from openpyxl import load_workbook

ROOT_DIR = Path(__file__).resolve().parents[2]
DEFAULT_XLSX = ROOT_DIR / "datosal" / "Ingredientes.xlsx"
DEFAULT_ENV = ROOT_DIR / "backend" / ".env"

def parse_env(path: Path) -> dict[str, str]:
    values: dict[str, str] = {}
    if not path.exists():
        return values
    for raw_line in path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        values[key.strip()] = value.strip().strip('"').strip("'")
    return values


def connect_pooler(url: str, host: str, port: int) -> psycopg.Connection:
    parsed = urlparse(url)
    project_ref = ""
    if parsed.hostname and parsed.hostname.startswith("db."):
        parts = parsed.hostname.split(".")
        if len(parts) > 1:
            project_ref = parts[1]

    user_candidates = []
    if project_ref:
        user_candidates.append(f"postgres.{project_ref}")
    if parsed.username and parsed.username not in user_candidates:
        user_candidates.append(parsed.username)
    if "postgres" not in user_candidates:
        user_candidates.append("postgres")

    dbname = parsed.path.lstrip("/") or "postgres"
    password = unquote(parsed.password or "")

    for user in user_candidates:
        try:
            return psycopg.connect(
                host=host,
                port=port,
                user=user,
                password=password,
                dbname=dbname,
                sslmode="require",
                connect_timeout=30,
                prepare_threshold=None,
            )
        except Exception:
            pass

    raise RuntimeError("No se pudo conectar")


def normalize_key(text: str) -> str:
    if not text:
        return ""
    no_accents = "".join(
        ch for ch in unicodedata.normalize("NFKD", str(text).strip())
        if not unicodedata.combining(ch)
    )
    return re.sub(r"\s+", " ", no_accents.lower().strip())


def strip_text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def to_float(value) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = strip_text(value)
    if not text:
        return None
    try:
        return float(text.replace(",", "."))
    except ValueError:
        return None


def load_rows(xlsx_path: Path) -> list[dict]:
    wb = load_workbook(xlsx_path, data_only=True)
    ws = None
    for name in wb.sheetnames:
        if normalize_key(name) == normalize_key("Composicion"):
            ws = wb[name]
            break

    if not ws:
        raise KeyError("No encontrada hoja Composicion")

    headers = []
    for col in range(1, ws.max_column + 1):
        headers.append(strip_text(ws.cell(1, col).value))

    rows = []
    for r in range(2, ws.max_row + 1):
        row = {}
        has_data = False
        for col, header in enumerate(headers, 1):
            val = ws.cell(r, col).value
            row[header] = val
            if val is not None:
                has_data = True
        if has_data and strip_text(row.get("Nombre")).strip():
            rows.append(row)

    return headers, rows


env = parse_env(DEFAULT_ENV)
url = env["DATABASE_URL"]

print("Conectando a Supabase...")
conn = connect_pooler(url, "aws-0-us-west-2.pooler.supabase.com", 6543)

headers, rows = load_rows(DEFAULT_XLSX)

print(f"Cargado {len(rows)} ingredientes, {len(headers)} columnas")

with conn.cursor() as cur:
    # Limpiar y cargar en lotes
    print("Insertandoingredientes en lotes de 50...")
    batch_size = 50
    for batch_start in range(0, len(rows), batch_size):
        batch = rows[batch_start : batch_start + batch_size]
        for row in batch:
            nombre = strip_text(row.get("Nombre"))
            if not nombre:
                continue

            grupo = strip_text(row.get("Grupo alimentario")) or "Sin clasificar"
            try:
                cur.execute(
                    """
                    insert into dom_nutricion_ingredientes.ingrediente (
                        nombre, nombre_ingles, id_grupo_alimentario, activo,
                        fuente_registro, fecha_importacion
                    ) values (
                        %s, %s,
                        (select id from dom_nutricion_catalogos.grupo_alimentario
                         where nombre = %s limit 1),
                        true, 'excel', now()
                    )
                    on conflict (nombre) do update set
                        fuente_registro = 'excel', fecha_importacion = now()
                    """,
                    (nombre, nombre, grupo),
                )
            except Exception as e:
                print(f"WARN fila: {nombre}: {e}")

        conn.commit()
        pct = min(100, int(100 * (batch_start + batch_size) / len(rows)))
        print(f"  {pct}% ({batch_start + len(batch)}/{len(rows)})")

print("CARGA COMPLETADA")
conn.close()
