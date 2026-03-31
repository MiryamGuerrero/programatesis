from __future__ import annotations

from pathlib import Path
from typing import Any
from urllib.parse import unquote, urlparse

import psycopg
from openpyxl import load_workbook

ROOT_DIR = Path(__file__).resolve().parents[2]
ENV_FILE = ROOT_DIR / "backend" / ".env"
DATA_DIR = ROOT_DIR / "datosal"

POOLER_REGIONS = [
    "us-east-1",
    "us-west-1",
    "us-west-2",
    "eu-west-1",
    "eu-west-2",
    "eu-central-1",
    "eu-north-1",
    "ap-southeast-1",
    "ap-southeast-2",
    "ap-northeast-1",
    "ap-south-1",
    "sa-east-1",
    "ca-central-1",
]

Z_SCORE_FILES = [
    "bmi-boys-z-who-2007-exp.xlsx",
    "bmi-girls-z-who-2007-exp (2).xlsx",
    "hfa-boys-z-who-2007-exp.xlsx",
    "hfa-girls-z-who-2007-exp (2).xlsx",
]

PERCENTILE_FILES = [
    "bmi-boys-perc-who2007-exp.xlsx",
    "bmi-girls-perc-who2007-exp.xlsx",
    "hfa-boys-perc-who2007-exp.xlsx",
    "hfa-girls-perc-who2007-exp.xlsx",
]

DDL_REFERENCIA = """
DROP TABLE IF EXISTS referencia.oms_referencia_percentil CASCADE;
DROP TABLE IF EXISTS referencia.oms_referencia_zscore CASCADE;
DROP TABLE IF EXISTS referencia.oms_percentil_referencia CASCADE;
DROP TABLE IF EXISTS referencia.oms_referencia CASCADE;
DROP TABLE IF EXISTS referencia.indicador_antropometrico CASCADE;

CREATE TABLE referencia.indicador_antropometrico (
    id SERIAL PRIMARY KEY,
    codigo VARCHAR(20) UNIQUE NOT NULL,
    nombre VARCHAR(150) NOT NULL,
    descripcion TEXT,
    activo BOOLEAN NOT NULL DEFAULT TRUE
);

INSERT INTO referencia.indicador_antropometrico (codigo, nombre, descripcion)
VALUES
    ('BMI', 'Indice de masa corporal para la edad', 'Body Mass Index-for-age (OMS 5-19 anos)'),
    ('HFA', 'Talla para la edad', 'Height-for-age (OMS 5-19 anos)')
ON CONFLICT (codigo) DO NOTHING;

CREATE TABLE referencia.oms_referencia_zscore (
    id BIGSERIAL PRIMARY KEY,
    id_indicador INTEGER NOT NULL REFERENCES referencia.indicador_antropometrico(id),
    id_sexo INTEGER NOT NULL REFERENCES usuarios.catalogo_sexo(id),
    edad_meses INTEGER NOT NULL CHECK (edad_meses BETWEEN 0 AND 240),
    l NUMERIC(12,6) NOT NULL,
    m NUMERIC(12,6) NOT NULL CHECK (m > 0),
    s NUMERIC(12,6) NOT NULL CHECK (s > 0),
    stdev NUMERIC(12,6),
    sd5neg NUMERIC(12,6),
    sd4neg NUMERIC(12,6) NOT NULL,
    sd3neg NUMERIC(12,6) NOT NULL,
    sd2neg NUMERIC(12,6) NOT NULL,
    sd1neg NUMERIC(12,6) NOT NULL,
    sd0 NUMERIC(12,6) NOT NULL,
    sd1 NUMERIC(12,6) NOT NULL,
    sd2 NUMERIC(12,6) NOT NULL,
    sd3 NUMERIC(12,6) NOT NULL,
    sd4 NUMERIC(12,6) NOT NULL,
    CHECK (sd5neg IS NULL OR sd5neg <= sd4neg),
    CHECK (
        sd4neg <= sd3neg AND
        sd3neg <= sd2neg AND
        sd2neg <= sd1neg AND
        sd1neg <= sd0 AND
        sd0 <= sd1 AND
        sd1 <= sd2 AND
        sd2 <= sd3 AND
        sd3 <= sd4
    ),
    UNIQUE(id_indicador, id_sexo, edad_meses)
);

CREATE INDEX idx_oms_referencia_zscore_sexo_edad
    ON referencia.oms_referencia_zscore(id_sexo, edad_meses);

CREATE TABLE referencia.oms_referencia_percentil (
    id BIGSERIAL PRIMARY KEY,
    id_indicador INTEGER NOT NULL REFERENCES referencia.indicador_antropometrico(id),
    id_sexo INTEGER NOT NULL REFERENCES usuarios.catalogo_sexo(id),
    edad_meses INTEGER NOT NULL CHECK (edad_meses BETWEEN 0 AND 240),
    l NUMERIC(12,6) NOT NULL,
    m NUMERIC(12,6) NOT NULL CHECK (m > 0),
    s NUMERIC(12,6) NOT NULL CHECK (s > 0),
    stdev NUMERIC(12,6),
    p01 NUMERIC(12,6) NOT NULL,
    p1 NUMERIC(12,6) NOT NULL,
    p3 NUMERIC(12,6) NOT NULL,
    p5 NUMERIC(12,6) NOT NULL,
    p10 NUMERIC(12,6) NOT NULL,
    p15 NUMERIC(12,6) NOT NULL,
    p25 NUMERIC(12,6) NOT NULL,
    p50 NUMERIC(12,6) NOT NULL,
    p75 NUMERIC(12,6) NOT NULL,
    p85 NUMERIC(12,6) NOT NULL,
    p90 NUMERIC(12,6) NOT NULL,
    p95 NUMERIC(12,6) NOT NULL,
    p97 NUMERIC(12,6) NOT NULL,
    p99 NUMERIC(12,6) NOT NULL,
    p999 NUMERIC(12,6) NOT NULL,
    CHECK (
        p01 <= p1 AND
        p1 <= p3 AND
        p3 <= p5 AND
        p5 <= p10 AND
        p10 <= p15 AND
        p15 <= p25 AND
        p25 <= p50 AND
        p50 <= p75 AND
        p75 <= p85 AND
        p85 <= p90 AND
        p90 <= p95 AND
        p95 <= p97 AND
        p97 <= p99 AND
        p99 <= p999
    ),
    UNIQUE(id_indicador, id_sexo, edad_meses)
);

CREATE INDEX idx_oms_referencia_percentil_sexo_edad
    ON referencia.oms_referencia_percentil(id_sexo, edad_meses);
"""


def parse_env_file(path: Path) -> dict[str, str]:
    values: dict[str, str] = {}
    if not path.exists():
        return values

    for raw_line in path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        values[key.strip()] = value.strip().strip('"').strip("'")

    return values


def normalize_header(value: Any) -> str:
    if value is None:
        return ""
    header = str(value).strip().lower()
    for token in [" ", "-", "+", ".", "_"]:
        header = header.replace(token, "")
    return header


def to_float(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, str):
        cleaned = value.strip().replace(",", ".")
        if not cleaned:
            return None
        try:
            return round(float(cleaned), 6)
        except ValueError:
            return None
    try:
        return round(float(value), 6)
    except (TypeError, ValueError):
        return None


def to_int_month(value: Any) -> int | None:
    parsed = to_float(value)
    if parsed is None:
        return None
    return int(round(parsed))


def detect_indicator(file_name: str) -> str:
    lower_name = file_name.lower()
    if "bmi" in lower_name:
        return "BMI"
    if "hfa" in lower_name:
        return "HFA"
    raise ValueError(f"No se pudo detectar indicador en {file_name}")


def detect_sex(file_name: str) -> str:
    lower_name = file_name.lower()
    if "boys" in lower_name:
        return "M"
    if "girls" in lower_name:
        return "F"
    raise ValueError(f"No se pudo detectar sexo en {file_name}")


def find_header_row(ws, required: set[str]) -> tuple[int, dict[str, int]]:
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=40, values_only=True), start=1):
        headers = [normalize_header(cell) for cell in row]
        mapping: dict[str, int] = {}
        for index, header in enumerate(headers):
            if header and header not in mapping:
                mapping[header] = index
        if required.issubset(mapping.keys()):
            return row_idx, mapping

    raise ValueError("No se encontro encabezado esperado")


def parse_zscore_file(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    worksheet = workbook[workbook.sheetnames[0]]

    required = {
        "month",
        "l",
        "m",
        "s",
        "sd4neg",
        "sd3neg",
        "sd2neg",
        "sd1neg",
        "sd0",
        "sd1",
        "sd2",
        "sd3",
        "sd4",
    }
    header_row, columns = find_header_row(worksheet, required)

    indicator_code = detect_indicator(path.name)
    sex_code = detect_sex(path.name)

    rows: list[dict[str, Any]] = []
    for row in worksheet.iter_rows(min_row=header_row + 1, values_only=True):
        edad_meses = to_int_month(row[columns["month"]])
        l = to_float(row[columns["l"]])
        m = to_float(row[columns["m"]])
        s = to_float(row[columns["s"]])

        sd4neg = to_float(row[columns["sd4neg"]])
        sd3neg = to_float(row[columns["sd3neg"]])
        sd2neg = to_float(row[columns["sd2neg"]])
        sd1neg = to_float(row[columns["sd1neg"]])
        sd0 = to_float(row[columns["sd0"]])
        sd1 = to_float(row[columns["sd1"]])
        sd2 = to_float(row[columns["sd2"]])
        sd3 = to_float(row[columns["sd3"]])
        sd4 = to_float(row[columns["sd4"]])

        if None in [edad_meses, l, m, s, sd4neg, sd3neg, sd2neg, sd1neg, sd0, sd1, sd2, sd3, sd4]:
            continue

        rows.append(
            {
                "indicador_codigo": indicator_code,
                "sexo_codigo": sex_code,
                "edad_meses": edad_meses,
                "l": l,
                "m": m,
                "s": s,
                "stdev": to_float(row[columns["stdev"]]) if "stdev" in columns else None,
                "sd5neg": to_float(row[columns["sd5neg"]]) if "sd5neg" in columns else None,
                "sd4neg": sd4neg,
                "sd3neg": sd3neg,
                "sd2neg": sd2neg,
                "sd1neg": sd1neg,
                "sd0": sd0,
                "sd1": sd1,
                "sd2": sd2,
                "sd3": sd3,
                "sd4": sd4,
            }
        )

    return rows


def parse_percentile_file(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    worksheet = workbook[workbook.sheetnames[0]]

    required = {
        "month",
        "l",
        "m",
        "s",
        "p01",
        "p1",
        "p3",
        "p5",
        "p10",
        "p15",
        "p25",
        "p50",
        "p75",
        "p85",
        "p90",
        "p95",
        "p97",
        "p99",
        "p999",
    }
    header_row, columns = find_header_row(worksheet, required)

    indicator_code = detect_indicator(path.name)
    sex_code = detect_sex(path.name)

    rows: list[dict[str, Any]] = []
    for row in worksheet.iter_rows(min_row=header_row + 1, values_only=True):
        edad_meses = to_int_month(row[columns["month"]])
        l = to_float(row[columns["l"]])
        m = to_float(row[columns["m"]])
        s = to_float(row[columns["s"]])

        p01 = to_float(row[columns["p01"]])
        p1 = to_float(row[columns["p1"]])
        p3 = to_float(row[columns["p3"]])
        p5 = to_float(row[columns["p5"]])
        p10 = to_float(row[columns["p10"]])
        p15 = to_float(row[columns["p15"]])
        p25 = to_float(row[columns["p25"]])
        p50 = to_float(row[columns["p50"]])
        p75 = to_float(row[columns["p75"]])
        p85 = to_float(row[columns["p85"]])
        p90 = to_float(row[columns["p90"]])
        p95 = to_float(row[columns["p95"]])
        p97 = to_float(row[columns["p97"]])
        p99 = to_float(row[columns["p99"]])
        p999 = to_float(row[columns["p999"]])

        if None in [
            edad_meses,
            l,
            m,
            s,
            p01,
            p1,
            p3,
            p5,
            p10,
            p15,
            p25,
            p50,
            p75,
            p85,
            p90,
            p95,
            p97,
            p99,
            p999,
        ]:
            continue

        rows.append(
            {
                "indicador_codigo": indicator_code,
                "sexo_codigo": sex_code,
                "edad_meses": edad_meses,
                "l": l,
                "m": m,
                "s": s,
                "stdev": to_float(row[columns["stdev"]]) if "stdev" in columns else None,
                "p01": p01,
                "p1": p1,
                "p3": p3,
                "p5": p5,
                "p10": p10,
                "p15": p15,
                "p25": p25,
                "p50": p50,
                "p75": p75,
                "p85": p85,
                "p90": p90,
                "p95": p95,
                "p97": p97,
                "p99": p99,
                "p999": p999,
            }
        )

    return rows


def resolve_ids(cursor) -> tuple[dict[str, int], dict[str, int]]:
    cursor.execute(
        """
        SELECT id, codigo
        FROM usuarios.catalogo_sexo
        WHERE codigo IN ('M', 'F')
        """
    )
    sex_ids = {codigo: sex_id for sex_id, codigo in cursor.fetchall()}
    if "M" not in sex_ids or "F" not in sex_ids:
        raise RuntimeError("No existen codigos M y F en usuarios.catalogo_sexo")

    cursor.execute(
        """
        SELECT id, codigo
        FROM referencia.indicador_antropometrico
        WHERE codigo IN ('BMI', 'HFA')
        """
    )
    indicator_ids = {codigo: indicator_id for indicator_id, codigo in cursor.fetchall()}
    if "BMI" not in indicator_ids or "HFA" not in indicator_ids:
        raise RuntimeError("No existen codigos BMI y HFA en referencia.indicador_antropometrico")

    return indicator_ids, sex_ids


def connect_with_fallback(database_url: str):
    try:
        return psycopg.connect(database_url)
    except Exception as direct_error:
        parsed = urlparse(database_url)
        host = parsed.hostname or ""
        dbname = parsed.path.lstrip("/") or "postgres"

        if not host.startswith("db.") or not host.endswith(".supabase.co"):
            raise RuntimeError(f"Conexion directa fallo y no se pudo inferir pooler: {direct_error}") from direct_error

        project_ref = host.split(".")[1]
        pooler_user = f"postgres.{project_ref}"
        pooler_password = unquote(parsed.password or "")

        last_error: Exception | None = None
        for region in POOLER_REGIONS:
            pooler_host = f"aws-0-{region}.pooler.supabase.com"
            try:
                connection = psycopg.connect(
                    host=pooler_host,
                    port=6543,
                    dbname=dbname,
                    user=pooler_user,
                    password=pooler_password,
                    sslmode="require",
                    connect_timeout=8,
                )
                print(f"Conexion via pooler habilitada: {pooler_host}")
                return connection
            except Exception as pooler_error:
                last_error = pooler_error

        raise RuntimeError(
            "No fue posible conectar por DB directa ni por pooler Supabase. "
            f"Ultimo error pooler: {last_error}"
        ) from direct_error


def main() -> None:
    env_values = parse_env_file(ENV_FILE)
    database_url = env_values.get("DATABASE_URL")
    if not database_url:
        raise RuntimeError("No se encontro DATABASE_URL en backend/.env")

    missing_files = [
        str(DATA_DIR / file_name)
        for file_name in [*Z_SCORE_FILES, *PERCENTILE_FILES]
        if not (DATA_DIR / file_name).exists()
    ]
    if missing_files:
        raise FileNotFoundError("Faltan archivos OMS requeridos:\n" + "\n".join(missing_files))

    z_rows: list[dict[str, Any]] = []
    for file_name in Z_SCORE_FILES:
        z_rows.extend(parse_zscore_file(DATA_DIR / file_name))

    p_rows: list[dict[str, Any]] = []
    for file_name in PERCENTILE_FILES:
        p_rows.extend(parse_percentile_file(DATA_DIR / file_name))

    with connect_with_fallback(database_url) as connection:
        with connection.cursor() as cursor:
            cursor.execute(DDL_REFERENCIA)
            indicator_ids, sex_ids = resolve_ids(cursor)

            z_payload = [
                (
                    indicator_ids[row["indicador_codigo"]],
                    sex_ids[row["sexo_codigo"]],
                    row["edad_meses"],
                    row["l"],
                    row["m"],
                    row["s"],
                    row["stdev"],
                    row["sd5neg"],
                    row["sd4neg"],
                    row["sd3neg"],
                    row["sd2neg"],
                    row["sd1neg"],
                    row["sd0"],
                    row["sd1"],
                    row["sd2"],
                    row["sd3"],
                    row["sd4"],
                )
                for row in z_rows
            ]

            cursor.executemany(
                """
                INSERT INTO referencia.oms_referencia_zscore (
                    id_indicador,
                    id_sexo,
                    edad_meses,
                    l,
                    m,
                    s,
                    stdev,
                    sd5neg,
                    sd4neg,
                    sd3neg,
                    sd2neg,
                    sd1neg,
                    sd0,
                    sd1,
                    sd2,
                    sd3,
                    sd4
                ) VALUES (
                    %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s
                )
                ON CONFLICT (id_indicador, id_sexo, edad_meses)
                DO UPDATE SET
                    l = EXCLUDED.l,
                    m = EXCLUDED.m,
                    s = EXCLUDED.s,
                    stdev = EXCLUDED.stdev,
                    sd5neg = EXCLUDED.sd5neg,
                    sd4neg = EXCLUDED.sd4neg,
                    sd3neg = EXCLUDED.sd3neg,
                    sd2neg = EXCLUDED.sd2neg,
                    sd1neg = EXCLUDED.sd1neg,
                    sd0 = EXCLUDED.sd0,
                    sd1 = EXCLUDED.sd1,
                    sd2 = EXCLUDED.sd2,
                    sd3 = EXCLUDED.sd3,
                    sd4 = EXCLUDED.sd4
                """,
                z_payload,
            )

            p_payload = [
                (
                    indicator_ids[row["indicador_codigo"]],
                    sex_ids[row["sexo_codigo"]],
                    row["edad_meses"],
                    row["l"],
                    row["m"],
                    row["s"],
                    row["stdev"],
                    row["p01"],
                    row["p1"],
                    row["p3"],
                    row["p5"],
                    row["p10"],
                    row["p15"],
                    row["p25"],
                    row["p50"],
                    row["p75"],
                    row["p85"],
                    row["p90"],
                    row["p95"],
                    row["p97"],
                    row["p99"],
                    row["p999"],
                )
                for row in p_rows
            ]

            cursor.executemany(
                """
                INSERT INTO referencia.oms_referencia_percentil (
                    id_indicador,
                    id_sexo,
                    edad_meses,
                    l,
                    m,
                    s,
                    stdev,
                    p01,
                    p1,
                    p3,
                    p5,
                    p10,
                    p15,
                    p25,
                    p50,
                    p75,
                    p85,
                    p90,
                    p95,
                    p97,
                    p99,
                    p999
                ) VALUES (
                    %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s
                )
                ON CONFLICT (id_indicador, id_sexo, edad_meses)
                DO UPDATE SET
                    l = EXCLUDED.l,
                    m = EXCLUDED.m,
                    s = EXCLUDED.s,
                    stdev = EXCLUDED.stdev,
                    p01 = EXCLUDED.p01,
                    p1 = EXCLUDED.p1,
                    p3 = EXCLUDED.p3,
                    p5 = EXCLUDED.p5,
                    p10 = EXCLUDED.p10,
                    p15 = EXCLUDED.p15,
                    p25 = EXCLUDED.p25,
                    p50 = EXCLUDED.p50,
                    p75 = EXCLUDED.p75,
                    p85 = EXCLUDED.p85,
                    p90 = EXCLUDED.p90,
                    p95 = EXCLUDED.p95,
                    p97 = EXCLUDED.p97,
                    p99 = EXCLUDED.p99,
                    p999 = EXCLUDED.p999
                """,
                p_payload,
            )

            cursor.execute(
                """
                SELECT i.codigo, s.codigo, COUNT(*)
                FROM referencia.oms_referencia_zscore t
                JOIN referencia.indicador_antropometrico i ON i.id = t.id_indicador
                JOIN usuarios.catalogo_sexo s ON s.id = t.id_sexo
                GROUP BY i.codigo, s.codigo
                ORDER BY i.codigo, s.codigo
                """
            )
            z_counts = cursor.fetchall()

            cursor.execute(
                """
                SELECT i.codigo, s.codigo, COUNT(*)
                FROM referencia.oms_referencia_percentil t
                JOIN referencia.indicador_antropometrico i ON i.id = t.id_indicador
                JOIN usuarios.catalogo_sexo s ON s.id = t.id_sexo
                GROUP BY i.codigo, s.codigo
                ORDER BY i.codigo, s.codigo
                """
            )
            p_counts = cursor.fetchall()

        connection.commit()

    print("Proceso completado correctamente.")
    print(f"Filas z-score procesadas: {len(z_rows)}")
    print(f"Filas percentil procesadas: {len(p_rows)}")
    print("Conteo z-score por indicador/sexo:")
    for indicator, sex, count in z_counts:
        print(f"- {indicator} / {sex}: {count}")
    print("Conteo percentil por indicador/sexo:")
    for indicator, sex, count in p_counts:
        print(f"- {indicator} / {sex}: {count}")


if __name__ == "__main__":
    main()
